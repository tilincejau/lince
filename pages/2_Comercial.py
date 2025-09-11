import streamlit as st
import pandas as pd
import re
import io
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# Setando a configuração da página para ampla
st.set_page_config(layout="wide")

# Verificação de login (simulada). Se não estiver logado, exibe uma mensagem e para o script.
if not st.session_state.get('is_logged_in'):
    st.session_state['is_logged_in'] = True # Simulando login para o script funcionar.
    # Em uma aplicação real, você teria uma lógica de autenticação aqui.

st.title("Setor Comercial")
st.markdown("Bem-vindo(a) ao setor Comercial. Abaixo estão os scripts disponíveis para análise.")

# --- Seleção do Script ---
script_selection = st.selectbox(
    "Selecione o script que deseja executar:",
    ("Análise de Canal", "Processador de Pontos")
)

# --- Seção 1: Análise de Canal e Validação de Dados ---
if script_selection == "Análise de Canal":
    st.write("---")
    st.subheader("📊 Análise de Canal")
    st.markdown("Este script transforma e consolida dados de planilhas de Google Forms, adicionando uma coluna de status com lista suspensa.")

    def normalize_columns(columns_list):
        """Normaliza uma lista de nomes de colunas."""
        normalized_list = []
        for col in columns_list:
            col = re.sub(r'\s+', ' ', col).strip()
            col = col.replace('\n', ' ')
            normalized_list.append(col)
        return normalized_list

    def transform_google_forms_data(df):
        """
        Transforma dados de Google Forms, consolidando informações e adicionando
        uma coluna 'Status' com validação de dados.
        """
        processed_records = []
        for index, row in df.iterrows():
            # Tratamento de dados baseado no script original
            data_value = row.iloc[0] if len(row) > 0 else None
            sv_value = row.iloc[1] if len(row) > 1 else None
            
            # Consolidar VD
            vd_consolidated_parts = [str(row.iloc[col_idx]).strip() for col_idx in range(2, min(5, len(row))) if pd.notna(row.iloc[col_idx])]
            vd_final = ' | '.join(vd_consolidated_parts) if vd_consolidated_parts else None
            
            # O valor do 'PARA' é a 28ª coluna (índice 27)
            para_value = row.iloc[27] if len(row) > 27 else None

            for col_idx in range(5, min(27, len(row))):
                cell_content = str(row.iloc[col_idx]).strip()
                if not cell_content or cell_content.lower() == 'nan':
                    continue
                
                de_category_match = re.search(r'\((.*?)\)', cell_content)
                de_category_val = de_category_match.group(1).strip() if de_category_match else None
                
                pdv_info_raw = re.sub(r'\s*\([^)]*\)\s*$', '', cell_content).strip()
                pdv_info_val = re.sub(r'^\s*(?:\b\w+\s+)?\d+\s*[\|-]\s*', '', pdv_info_raw, 1).strip() if pdv_info_raw else None
                
                if pdv_info_val or de_category_val:
                    processed_records.append({
                        'DATA': data_value,
                        'SV': sv_value,
                        'VD': vd_final,
                        'PDV': pdv_info_val,
                        'DE': de_category_val,
                        'PARA': para_value,
                        'Status': '' 
                    })

        final_df = pd.DataFrame(processed_records)
        return final_df

    # Campo de texto para as opções da lista suspensa
    dropdown_input = st.text_input(
        "Insira as opções da lista suspensa (separadas por vírgula):",
        "Aprovado,Não Aprovado"
    )

    uploaded_file_1 = st.file_uploader("Envie o arquivo para 'Análise de Canal' (.xlsx)", type=["xlsx"])

    if uploaded_file_1 is not None:
        try:
            df_forms = pd.read_excel(uploaded_file_1)
            st.subheader("📄 Dados Originais (Análise de Canal)")
            st.dataframe(df_forms.head())
            
            final_df_forms = transform_google_forms_data(df_forms)
            
            output = io.BytesIO()
            final_df_forms.to_excel(output, index=False)
            output.seek(0)
            
            workbook = load_workbook(output)
            sheet = workbook.active
            
            # Usa as opções do campo de texto para a validação de dados
            dropdown_options_excel = f'"{dropdown_input}"'
            dv = DataValidation(type="list", formula1=dropdown_options_excel, allow_blank=True)
            dv.error = 'O valor inserido não está na lista.'
            dv.errorTitle = 'Valor Inválido'
            
            try:
                col_for_dropdown_letter = get_column_letter(final_df_forms.columns.get_loc('Status') + 1)
                dv.add(f'{col_for_dropdown_letter}2:{col_for_dropdown_letter}{sheet.max_row}')
                sheet.add_data_validation(dv)
            except KeyError:
                st.warning("A coluna 'Status' não foi encontrada no DataFrame final.")
            
            output_with_dropdown = io.BytesIO()
            workbook.save(output_with_dropdown)
            output_with_dropdown.seek(0)
            
            st.subheader("✅ Dados Transformados (Análise de Canal)")
            st.dataframe(final_df_forms)
            
            st.download_button(
                label="📥 Baixar Arquivo de Análise de Canal",
                data=output_with_dropdown.getvalue(),
                file_name="analise_canal_processada.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
        except Exception as e:
            st.error(f"Ocorreu um erro durante o processamento de 'Análise de Canal': {e}")

# --- Seção 2: Processador de Arquivos para Pontuação ---
elif script_selection == "Processador de Pontos":
    st.write("---")
    st.subheader("⚙️ Processador de Arquivos")
    st.markdown("Este script converte os valores 'Presença' em pontuação, com base no nome das colunas.")

    def extract_points(column_name):
        """Função para extrair o valor numérico entre parênteses em uma string de cabeçalho."""
        match = re.search(r"\(\s*(\d+)\s*Pontos\s*\)", column_name)
        return int(match.group(1)) if match else None

    def transform_points_columns(df):
        """
        Aplica a transformação de 'Presença' para pontos nas colunas
        que contêm 'Pontos' no nome.
        """
        df_transformed = df.copy()
        for col in df_transformed.columns:
            if "Pontos" in col:
                points = extract_points(col)
                if points is not None:
                    # Usa 'apply' para substituir "Presença" pelo valor de pontos e outros por 0
                    df_transformed[col] = df_transformed[col].apply(lambda x: points if x == "Presença" else 0)
        return df_transformed

    @st.cache_data
    def convert_df_to_excel(df):
        """Converte DataFrame para um arquivo Excel em memória."""
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            df.to_excel(writer, index=False)
        processed_data = output.getvalue()
        return processed_data

    uploaded_file_2 = st.file_uploader("Envie o arquivo para 'Processador de Pontos' (.xlsx)", type=["xlsx"])

    if uploaded_file_2 is not None:
        try:
            df_points = pd.read_excel(uploaded_file_2)
            st.subheader("📄 Dados Originais (Processador de Pontos)")
            st.dataframe(df_points)
            
            df_transformed_points = transform_points_columns(df_points)
            
            st.subheader("✅ Dados Transformados (Processador de Pontos)")
            st.dataframe(df_transformed_points)
            
            excel_data = convert_df_to_excel(df_transformed_points)
            
            st.download_button(
                label="📥 Baixar Arquivo de Pontos Transformado",
                data=excel_data,
                file_name="pontos_transformados.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
        except Exception as e:
            st.error(f"Ocorreu um erro durante o processamento de 'Processador de Pontos': {e}")
