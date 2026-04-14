import streamlit as st
import pandas as pd
import re
import io
import numpy as np
from datetime import datetime, timedelta
import PyPDF2
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import xlsxwriter

# --- BIBLIOTECAS PARA GOOGLE SHEETS ---
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from gspread_dataframe import set_with_dataframe, get_as_dataframe

# ====================================================================
# 1. CONFIGURAÇÃO DA PÁGINA
# ====================================================================
st.set_page_config(
    page_title="Lince Distribuidora - Nuvem", 
    page_icon="☁️", 
    layout="centered"
)

st.markdown("""
<style>
    .stApp { background-color: #f0f2f6; } 
    div.stButton > button:first-child { background-color: #007bff; color: white; border-radius: 5px; font-weight: bold;} 
    .stTitle { text-align: center; color: #004d99; font-family: 'Arial', sans-serif;}
    h1, h2, h3 { color: #004d99; }
</style>
""", unsafe_allow_html=True)

# ====================================================================
# 2. CONFIGURAÇÃO E CONSTANTES GLOBAIS
# ====================================================================

SPREADSHEET_KEY = '1uFr9yhylYj7dINsDAr-6tECgNDxc21t9QhmC0cxBjhY' 

NAME_540_001 = '540-001 - GARRAFA 600ML' 
NAME_550_001 = '550-001 - CAIXA PLASTICA 600ML'

CRATE_TO_BOTTLE_MAP = {
    '546-004 - CAIXA PLASTICA 24UN 300ML': '546-001 - GARRAFA 300ML',
    '550-001 - CAIXA PLASTICA 600ML': NAME_540_001, 
    '587-002 - CAIXA PLASTICA HEINEKEN 600ML': '586-001 - GARRAFA HEINEKEN 600ML',
    '591-002 - CAIXA PLASTICA HEINEKEN 330ML': '593-001 - GARRAFA HEINEKEN 330ML',
    '555-001 - CAIXA PLASTICA 1L': '541-002 - GARRAFA 1L'
}

FACTORS = {
    '546-004 - CAIXA PLASTICA 24UN 300ML': 24,
    '550-001 - CAIXA PLASTICA 600ML': 24,
    '587-002 - CAIXA PLASTICA HEINEKEN 600ML': 24,
    '591-002 - CAIXA PLASTICA HEINEKEN 330ML': 24,
    '555-001 - CAIXA PLASTICA 1L': 12
}

# ====================================================================
# 3. CONEXÃO GOOGLE SHEETS
# ====================================================================

@st.cache_resource
def connect_to_gsheets():
    """Conecta ao Google Sheets usando Streamlit Secrets (Nuvem) ou arquivo local"""
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    
    try:
        if "gcp_service_account" in st.secrets:
            creds_dict = dict(st.secrets["gcp_service_account"])
            creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
        else:
            creds = ServiceAccountCredentials.from_json_keyfile_name("credentials.json", scope)
            
        client = gspread.authorize(creds)
        
        try:
            sheet = client.open_by_key(SPREADSHEET_KEY)
            return sheet
        except gspread.SpreadsheetNotFound:
            st.error("Planilha não encontrada! Verifique o ID e se você compartilhou com o email do bot.")
            return None
    except Exception as e:
        st.error(f"Erro na autenticação do Google: {e}")
        return None

def load_from_gsheets(sheet, tab_name):
    """Lê uma aba específica da planilha e retorna como DataFrame"""
    try:
        try:
            worksheet = sheet.worksheet(tab_name)
        except gspread.WorksheetNotFound:
            return pd.DataFrame() 

        df = get_as_dataframe(worksheet, evaluate_formulas=True, dtype=str) 
        df = df.dropna(how='all').dropna(axis=1, how='all')

        cols_date = ['DataCompleta', 'DataCompleta_excel', 'DataCompleta_txt', 'DataCompleta_pdf']
        for col in cols_date:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors='coerce')
        
        for col in df.columns:
            if col not in cols_date and col != 'Vasilhame' and col != 'Dia':
                df[col] = pd.to_numeric(df[col], errors='ignore')

        return df
    except Exception as e:
        st.warning(f"Erro ao ler aba {tab_name}: {e}")
        return pd.DataFrame()

def save_to_gsheets(sheet, tab_name, df):
    """Salva o DataFrame em uma aba, sobrescrevendo ou criando"""
    try:
        try:
            worksheet = sheet.worksheet(tab_name)
            worksheet.clear() 
        except gspread.WorksheetNotFound:
            worksheet = sheet.add_worksheet(title=tab_name, rows="1000", cols="20")
        
        df_export = df.copy()
        for col in df_export.select_dtypes(include=['datetime64[ns]']).columns:
             df_export[col] = df_export[col].astype(str).replace('NaT', '')

        set_with_dataframe(worksheet, df_export)
        return True
    except Exception as e:
        st.error(f"Erro ao salvar na aba {tab_name}: {e}")
        return False

# ====================================================================
# 4. SISTEMA DE LOGIN
# ====================================================================

def login_form():
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("<h2 style='text-align: center;'>Lince Distribuidora</h2>", unsafe_allow_html=True)
        st.markdown("---")
        with st.form("login_form", clear_on_submit=False):
            st.markdown("Por favor, insira suas credenciais.")
            username = st.text_input("Usuário", key="username_input", placeholder="Digite seu usuário")
            password = st.text_input("Senha", type="password", key="password_input", placeholder="Digite sua senha")
            st.markdown("<br>", unsafe_allow_html=True)
            submit_button = st.form_submit_button("Entrar", use_container_width=True)
        
        if submit_button:
            if username in st.session_state.LOGIN_INFO and st.session_state.LOGIN_INFO[username] == password:
                st.session_state['is_logged_in'] = True
                st.session_state['username'] = username
                st.session_state['current_page'] = 'home'
                st.rerun()
            else:
                st.error("Usuário ou senha incorretos.")

def main_page():
    st.markdown(f"<h1 style='text-align: center;'>Página Inicial</h1>", unsafe_allow_html=True)
    st.markdown(f"<p style='text-align: center;'>Bem-vindo(a), <b>{st.session_state['username']}</b>!</p>", unsafe_allow_html=True)
    st.markdown("---")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        if st.button("🚛 Logística", use_container_width=True):
            st.session_state['current_page'] = 'logistics'
            st.rerun()
    with col2:
        if st.button("📈 Comercial", use_container_width=True):
            st.session_state['current_page'] = 'commercial'
            st.rerun()
    with col3:
        if st.button("📊 Assessment", use_container_width=True):
            st.session_state['current_page'] = 'assessment'
            st.rerun()
            
    st.markdown("---")
    if st.button("Sair", use_container_width=True):
        st.session_state['is_logged_in'] = False
        st.session_state.pop('username', None)
        st.session_state.pop('current_page', None)
        st.rerun()
# ====================================================================
# 5. SETOR DE LOGÍSTICA
# ====================================================================
def logistics_page():
    st.title("Setor de Logística")
    
    col_voltar, col_vazio = st.columns([1, 5])
    with col_voltar:
        if st.button("⬅️ Voltar"):
            st.session_state['current_page'] = 'home'
            st.rerun()

    st.markdown("---")
    
    script_choice = st.selectbox(
        "Selecione uma ferramenta:",
        ("Selecione...", "Acurácia", "Validade", "Vasilhames", "Abastecimento", "Manutenção Veículos"),
        key="log_select" 
    )
    
    st.write("---")

    # --- SCRIPT ACURÁCIA ---
    if script_choice == "Acurácia":
        st.subheader("Acurácia de Estoque")
        uploaded_file = st.file_uploader("Envie o arquivo 'Acuracia estoque' (.csv ou .xlsx)", type=["csv", "xlsx"], key="acuracia_uploader")
        if uploaded_file is not None:
            try:
                df = None
                if uploaded_file.name.endswith('.csv'):
                    df = pd.read_csv(uploaded_file, header=[0, 1])
                elif uploaded_file.name.endswith('.xlsx'):
                    df = pd.read_excel(uploaded_file, header=[0, 1], sheet_name=0)
                else:
                    st.error("Formato de arquivo não suportado."); return 
                products_to_remove = ['185039 - Garrafa 0,30l', '471 - Garrafa 0,60l (3 )']
                try:
                    prod_cod_col = df.columns[0]
                    df_data = df.set_index(prod_cod_col)
                except IndexError: st.error("Erro ao definir o índice do DataFrame."); return
                df_data = df_data[~df_data.index.isin(products_to_remove)].copy()
                df_data = df_data[~df_data.index.astype(str).str.contains('Totais', na=False)].copy()
                data_types_from_file = ['Contagem - $', 'Diferença - $', 'Saldo Final - $'] 
                first_level_cols = [col[0] for col in df.columns]
                unique_dates = sorted(list(set([col for col in first_level_cols if col not in ['Data', 'Prod Cód', 'Totais'] and 'Unnamed' not in str(col)])))
                new_rows = []
                for product in df_data.index:
                    for date in unique_dates:
                        row_data = {'Prod Cód': product, 'Dia': date}
                        for data_type in data_types_from_file: 
                            try:
                                col_name = (date, data_type)
                                value = df_data.loc[product, col_name]
                                if isinstance(value, str):
                                    if value.strip() == '-': value = 0
                                row_data[data_type] = pd.to_numeric(value, errors='coerce')
                            except KeyError: row_data[data_type] = np.nan
                        new_rows.append(row_data)
                df_final = pd.DataFrame(new_rows)
                df_final.rename(columns={'Contagem - $': 'Contagem', 'Diferença - $': 'Diferença', 'Saldo Final - $': 'Saldo Final'}, inplace=True)
                df_final['Saldo Final'] = df_final['Saldo Final'].fillna(0).apply(lambda x: max(0, x))
                df_final['Diferença'] = df_final['Diferença'].fillna(0).abs()
                df_final['Contagem'] = df_final['Contagem'].fillna(0)
                df_final = df_final.sort_values(by=['Dia', 'Prod Cód'])
                df_final['Dia'] = pd.to_datetime(df_final['Dia']).dt.strftime('%Y-%m-%d')
                numeric_cols = ['Saldo Final', 'Contagem', 'Diferença'] 
                existing_numeric_cols = [col for col in numeric_cols if col in df_final.columns]
                df_final[existing_numeric_cols] = df_final[existing_numeric_cols].round(2)
                desired_order = ['Prod Cód', 'Dia', 'Contagem', 'Diferença', 'Saldo Final']
                df_final = df_final[desired_order]
                st.subheader("📊 Resultado da Acurácia")
                st.dataframe(df_final)
                excel_data = io.BytesIO()
                df_final.to_excel(excel_data, index=False, engine='xlsxwriter')
                excel_data.seek(0)
                st.download_button(label="📥 Baixar Arquivo Processado", data=excel_data, file_name='Acuracia_estoque_processado.xlsx', mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            except Exception as e:
                st.error(f"Ocorreu um erro no script de Acurácia: {e}")

    # --- SCRIPT VALIDADE ---
    elif script_choice == "Validade":
        st.subheader("Controle de Validade")

        def parse_estoque_txt_st(file_content):
            lines = [line.decode('latin1') for line in file_content.getvalue().splitlines()]
            separator_string = '-' * 116
            separator_indices = [i for i, line in enumerate(lines) if separator_string in line]
            if len(separator_indices) < 2: return pd.DataFrame()
            start_line = separator_indices[1] + 1
            col_names = ['COD.RED.', 'DESCRIÇÃO', 'SLD INICIAL CX', 'SLD INICIAL UN', 'ENTRADAS CX', 'ENTRADAS UN', 'SAÍDAS CX', 'SAÍDAS UN', 'SALDO FÍSICO CX', 'SALDO FÍSICO UN', 'CONT. FÍSICA CX', 'CONT. FÍSICA UN', 'DIFERENÇA CX', 'DIFERENÇA UN']
            data = []
            pattern = re.compile(r'^\s*(\d+)\s+(.+?)\s*([-+]?\d*)\s*([-+]?\d*)\s*I\s*([-+]?\d*)\s*([-+]?\d*)\s*I\s*([-+]?\d*)\s*([-+]?\d*)\s*I\s*([-+]?\d*)\s*([-+]?\d*)\s*I\s*([-+]?\d*)\s*([-+]?\d*)\s*I\s*([-+]?\d*)\s*([-+]?\d*)\s*I')
            for line in lines[start_line:]:
                line = line.strip()
                if not line or 'TOTAL GERAL' in line: continue
                match = pattern.match(line)
                if match:
                    groups = list(match.groups())
                    row_values = [groups[0], groups[1].strip()]
                    for i in range(2, len(groups), 2):
                        cx = groups[i].strip() if groups[i] and groups[i].strip() else '0'
                        un = groups[i+1].strip() if groups[i+1] and groups[i+1].strip() else '0'
                        row_values.extend([int(cx), int(un)])
                    if len(row_values) == 14: data.append(row_values)
            return pd.DataFrame(data, columns=col_names)

        def extract_units_per_box(product_name):
            product_name = str(product_name).upper().replace(' ', '')
            match_multiplication = re.search(r'(\d+)X(\d+)(?:UN|U)', product_name)
            if match_multiplication: return int(match_multiplication.group(1)) * int(match_multiplication.group(2))
            match_direct = re.search(r'(\d+)(?:UN|U)', product_name)
            if match_direct: return int(match_direct.group(1)) 
            return 1

        uploaded_excel_file = st.file_uploader("Envie o arquivo Excel 'Controle de Validade.xlsx'", type=["xlsx"], key="validade_excel_uploader") 
        uploaded_txt_file = st.file_uploader("Envie o arquivo de texto de estoque", type=["txt"], key="validade_txt_uploader")

        if uploaded_excel_file is not None and uploaded_txt_file is not None:
            try:
                df_validade = pd.read_excel(uploaded_excel_file)
                df_validade.columns = df_validade.columns.str.replace(r'\s+', ' ', regex=True).str.strip()

                df_estoque = parse_estoque_txt_st(uploaded_txt_file)
                if df_estoque.empty: st.warning("O arquivo TXT está vazio ou não pôde ser processado."); return

                validity_cols = ['Validade', 'Validade 2', 'Validade 3', 'Validade 4', 'Validade 5']
                quantity_caixa_cols = ['Quantidade (CAIXA)', 'Quantidade 2 (CAIXA)', 'Quantidade 3 (CAIXA)', 'Quantidade 4 (CAIXA)', 'Quantidade 5 (CAIXA)']
                quantity_unidade_cols = ['Quantidade (UNIDADE)', 'Quantidade 2 (UNIDADE)', 'Quantidade 3 (UNIDADE)', 'Quantidade 4 (UNIDADE)', 'Quantidade 5 (UNIDADE)']
                
                all_validity_entries = []
                
                for i in range(len(validity_cols)):
                    v_col = validity_cols[i]
                    c_col = quantity_caixa_cols[i]
                    u_col = quantity_unidade_cols[i]

                    if v_col not in df_validade.columns:
                        v_col_alt = v_col.replace(' ', '.') 
                        if v_col_alt in df_validade.columns:
                            v_col = v_col_alt
                    
                    cols_to_check = ['Qual Produto ?', v_col]
                    
                    if all(col in df_validade.columns for col in cols_to_check):
                        cols_select = ['Qual Produto ?', v_col]
                        if c_col in df_validade.columns: cols_select.append(c_col)
                        if u_col in df_validade.columns: cols_select.append(u_col)

                        temp_df = df_validade[cols_select].copy()
                        
                        rename_map = {v_col: 'Validade'}
                        if c_col in temp_df.columns: rename_map[c_col] = 'Quantidade (CAIXA)'
                        if u_col in temp_df.columns: rename_map[u_col] = 'Quantidade (UNIDADE)'
                        
                        temp_df.rename(columns=rename_map, inplace=True)
                        
                        if 'Quantidade (CAIXA)' not in temp_df.columns: temp_df['Quantidade (CAIXA)'] = 0
                        if 'Quantidade (UNIDADE)' not in temp_df.columns: temp_df['Quantidade (UNIDADE)'] = 0

                        all_validity_entries.append(temp_df)

                all_validity_entries = [df for df in all_validity_entries if not df.dropna(subset=['Validade']).empty]
                
                melted_df_validade_all = pd.concat(all_validity_entries, ignore_index=True) if all_validity_entries else pd.DataFrame(columns=['Qual Produto ?', 'Validade', 'Quantidade (CAIXA)', 'Quantidade (UNIDADE)'])
                
                melted_df_validade_all.dropna(subset=['Validade'], inplace=True)
                melted_df_validade_all['Validade'] = pd.to_datetime(melted_df_validade_all['Validade'], errors='coerce')
                melted_df_validade_all.dropna(subset=['Validade'], inplace=True)
                
                melted_df_validade_all['Quantidade (CAIXA)'] = pd.to_numeric(melted_df_validade_all['Quantidade (CAIXA)'], errors='coerce').fillna(0)
                melted_df_validade_all['Quantidade (UNIDADE)'] = pd.to_numeric(melted_df_validade_all['Quantidade (UNIDADE)'], errors='coerce').fillna(0)
                
                split_data_validade = melted_df_validade_all['Qual Produto ?'].astype(str).str.split(' - ', n=1, expand=True)
                melted_df_validade_all['Codigo Produto'] = split_data_validade[0].str.strip()
                melted_df_validade_all['Nome Produto'] = split_data_validade[1].str.strip()
                
                melted_df_validade_all['Units_Per_Box_Temp'] = melted_df_validade_all['Nome Produto'].apply(extract_units_per_box)
                
                grouped = melted_df_validade_all.groupby(['Codigo Produto', 'Nome Produto', 'Validade']).agg({'Quantidade (CAIXA)': 'sum', 'Quantidade (UNIDADE)': 'sum', 'Units_Per_Box_Temp': 'first'}).reset_index()
                
                def convert_total_units_to_boxes_and_units(row):
                    units_per_box = row['Units_Per_Box_Temp'] or 1
                    total_units = (row['Quantidade (CAIXA)'] * units_per_box) + row['Quantidade (UNIDADE)']
                    row['Quantidade (CAIXA)'] = total_units // units_per_box
                    row['Quantidade (UNIDADE)'] = total_units % units_per_box
                    return row
                    
                grouped = grouped.apply(convert_total_units_to_boxes_and_units, axis=1)
                grouped.drop('Units_Per_Box_Temp', axis=1, inplace=True)
                
                data_atual = datetime.now()
                grouped['Dias para Vencer'] = (grouped['Validade'] - data_atual).dt.days
                
                conditions = [grouped['Dias para Vencer'] <= 45, (grouped['Dias para Vencer'] > 45) & (grouped['Dias para Vencer'] <= 60), grouped['Dias para Vencer'] > 60]
                choices = ['VALIDADE CURTA', 'ATENÇÃO', 'OK']
                grouped['Status Validade'] = np.select(conditions, choices, default='Indefinido')
                
                grouped['Validade_DateOnly'] = grouped['Validade'].dt.date
                
                sorted_grouped = grouped.sort_values(by=['Codigo Produto', 'Validade']).reset_index(drop=True)
                sorted_grouped['Validade_Rank'] = sorted_grouped.groupby('Codigo Produto')['Validade'].rank(method='first').astype(int)
                
                final_rows = []
                for product_code, group in sorted_grouped.groupby('Codigo Produto'):
                    row = {'Codigo Produto': product_code, 'Nome Produto': group['Nome Produto'].iloc[0]}
                    for _, r in group.iterrows():
                        i = r['Validade_Rank']
                        row[f'Validade {i}'] = r['Validade_DateOnly']
                        row[f'Quantidade (CAIXA) {i}'] = r['Quantidade (CAIXA)']
                        row[f'Quantidade (UNIDADE) {i}'] = r['Quantidade (UNIDADE)']
                        row[f'Dias para Vencer {i}'] = r['Dias para Vencer']
                        row[f'Status Validade {i}'] = r['Status Validade']
                    final_rows.append(row)
                    
                final_df = pd.DataFrame(final_rows)
                
                if not df_estoque.empty:
                    df_estoque['COD.RED.'] = df_estoque['COD.RED.'].astype(str)
                    final_df['Codigo Produto'] = final_df['Codigo Produto'].astype(str)
                    
                    df_saldo = df_estoque[['COD.RED.', 'SALDO FÍSICO CX', 'SALDO FÍSICO UN']].drop_duplicates('COD.RED.')
                    df_saldo.rename(columns={'SALDO FÍSICO CX': 'Saldo Físico TXT Caixa', 'SALDO FÍSICO UN': 'Saldo Físico TXT Unidade'}, inplace=True)
                    
                    final_df = pd.merge(final_df, df_saldo, how='left', left_on='Codigo Produto', right_on='COD.RED.')
                    final_df.drop('COD.RED.', axis=1, inplace=True)
                
                quantidade_caixa_cols = [col for col in final_df.columns if re.match(r'Quantidade \(CAIXA\) \d+', col)]
                quantidade_unidade_cols = [col for col in final_df.columns if re.match(r'Quantidade \(UNIDADE\) \d+', col)]
                
                final_df['Contagem Fisica CX'] = final_df[quantidade_caixa_cols].sum(axis=1)
                final_df['Contagem Fisica UN'] = final_df[quantidade_unidade_cols].sum(axis=1)
                
                st.subheader("✅ Relatório de Validade Gerado")
                st.dataframe(final_df)
                
                excel_data = io.BytesIO()
                final_df.to_excel(excel_data, sheet_name='Controle de Estoque', index=False)
                excel_data.seek(0)
                st.download_button(label="📥 Baixar Relatório de Validade", data=excel_data, file_name="Controle_Estoque_Completo.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                
            except Exception as e:
                st.error(f"Ocorreu um erro ao processar os arquivos: {e}")

    # --- SCRIPT VASILHAMES ---
    elif script_choice == "Vasilhames":
        st.subheader("Controle de Vasilhames (Nuvem ☁️)")
        
        sheet_client = connect_to_gsheets()
        
        if not sheet_client:
            st.error("Não foi possível conectar ao Google Sheets. Verifique o arquivo credentials.json e o ID da planilha.")
            st.stop()

        st.write("---")
        st.subheader("⚙️ Gerenciamento")
        
        with st.expander("🔴 ZONA DE PERIGO: Limpar Banco de Dados (Clique para abrir)"):
            st.warning("⚠️ ATENÇÃO: Esta ação é IRREVERSÍVEL!")
            st.markdown("Ao clicar no botão abaixo, **todo o histórico** salvo nas planilhas do Google (TXT, PDF, Vendas, Excel) será apagado permanentemente.")
            
            trava_seguranca = st.checkbox("Sim, eu tenho certeza e quero apagar todo o histórico.")
            
            if trava_seguranca:
                if st.button("🗑️ CONFIRMAR LIMPEZA TOTAL", type="primary"):
                    try:
                        abas_para_limpar = ['txt_data', 'pdf_data', 'vendas_data', 'excel_data', 'CONSOLIDADO_GERAL']
                        bar = st.progress(0)
                        
                        for i, tab in enumerate(abas_para_limpar):
                            try:
                                ws = sheet_client.worksheet(tab)
                                ws.clear()
                            except: pass
                            bar.progress((i + 1) / len(abas_para_limpar))
                            
                        st.success("Histórico na nuvem apagado com sucesso!")
                        import time
                        time.sleep(2)
                        st.rerun()
                    except Exception as e:
                        st.error(f"Erro ao limpar o banco: {e}")
            else:
                st.info("Para liberar o botão de apagar, marque a caixa de confirmação acima.")

        st.write("---")
    
        def process_vendas_file(file_content):
            content = file_content.getvalue().decode('latin1')
            filename_date_match = re.search(r'VENDA(\d{4})\.TXT', file_content.name)
            effective_date_str = None
            effective_date_full = None
            
            if filename_date_match:
                day = filename_date_match.group(1)[:2]
                month = filename_date_match.group(1)[2:]
                year = datetime.now().year
                now = datetime.now()
                if now.month == 1 and month == '12': year = year - 1
                effective_date_obj = datetime.strptime(f"{day}/{month}/{year}", '%d/%m/%Y')
                effective_date_str = effective_date_obj.strftime('%d/%m')
                effective_date_full = effective_date_obj.date()
            else:
                 effective_date_obj = datetime.now()
                 effective_date_str = effective_date_obj.strftime('%d/%m')
                 effective_date_full = effective_date_obj.date()

            sales_map = {
                '540-001': NAME_540_001,
                '541-002': '541-002 - GARRAFA 1L',
                '586-001': '586-001 - GARRAFA HEINEKEN 600ML',
                '593-001': '593-001 - GARRAFA HEINEKEN 330ML',
                '555-001': '555-001 - CAIXA PLASTICA 1L',
                '587-002': '587-002 - CAIXA PLASTICA HEINEKEN 600ML',
                '591-002': '591-002 - CAIXA PLASTICA HEINEKEN 330ML',
                '803-039': NAME_550_001, 
                '550-001': NAME_550_001, 
                '550-012': NAME_550_001 
            }

            parsed_data = []
            lines = content.splitlines()
            
            for line in lines:
                line = line.strip()
                match = re.search(r'^(\d{6}).*?([\d\.]+)\s*\/', line)
                
                if match:
                    raw_code = match.group(1)
                    raw_qty = match.group(2)
                    qty = int(raw_qty.replace('.', ''))
                    formatted_code = f"{raw_code[:3]}-{raw_code[3:]}"
                    
                    if formatted_code in sales_map:
                        vasilhame = sales_map[formatted_code]
                        parsed_data.append({'Vasilhame': vasilhame, 'Vendas': qty, 'Dia': effective_date_str, 'DataCompleta': effective_date_full})

            if not parsed_data: return None
            return pd.DataFrame(parsed_data)

        def process_txt_file_st(file_content):
            content = file_content.getvalue().decode('latin1')
            filename_date_match = re.search(r'ESTOQUE(\d{4})\.TXT', file_content.name)
            effective_date_str = None
            effective_date_full = None
            if filename_date_match:
                day = filename_date_match.group(1)[:2]
                month = filename_date_match.group(1)[2:]
                year = datetime.now().year
                now = datetime.now()
                if now.month == 1 and month == '12': year = year - 1
                effective_date_obj = datetime.strptime(f"{day}/{month}/{year}", '%d/%m/%Y')
                effective_date_str = effective_date_obj.strftime('%d/%m')
                effective_date_full = effective_date_obj.date()
            else: st.error(f"Nome do arquivo TXT inválido: {file_content.name}"); return None, None, None 

            product_code_to_vasilhame_map = {
                '563-008': '563-008 - BARRIL INOX 30L', '564-009': '564-009 - BARRIL INOX 50L', '591-002': '591-002 - CAIXA PLASTICA HEINEKEN 330ML', 
                '587-002': '587-002 - CAIXA PLASTICA HEINEKEN 600ML', '550-001': '550-001 - CAIXA PLASTICA 600ML', '555-001': '555-001 - CAIXA PLASTICA 1L', 
                '546-004': '546-004 - CAIXA PLASTICA 24UN 300ML', '565-002': '565-002 - CILINDRO CO2', 
                '546-001': CRATE_TO_BOTTLE_MAP['546-004 - CAIXA PLASTICA 24UN 300ML'],
                '540-001': NAME_540_001, 
                '541-002': CRATE_TO_BOTTLE_MAP['555-001 - CAIXA PLASTICA 1L'],
                '586-001': CRATE_TO_BOTTLE_MAP['587-002 - CAIXA PLASTICA HEINEKEN 600ML'],
                '593-001': CRATE_TO_BOTTLE_MAP['591-002 - CAIXA PLASTICA HEINEKEN 330ML'],
                '550-012': '550-001 - CAIXA PLASTICA 600ML', '803-025': '550-001 - CAIXA PLASTICA 600ML',
                '803-036': '550-001 - CAIXA PLASTICA 600ML', '803-037': '550-001 - CAIXA PLASTICA 600ML',
                '803-039': '550-001 - CAIXA PLASTICA 600ML' 
            }

            parsed_data = []
            lines = content.splitlines()
            current_code = None
            for line in lines:
                line = line.strip()
                if not line or '---' in line or 'DATA' in line or 'REFERENTE' in line: continue
                code_match = re.search(r'^(\d{3}-\d{3})', line)
                if code_match:
                    current_code = code_match.group(1)
                    qty_match_full = re.search(r'\s+([\d\.]+)\s+[\d\.]+,\d+', line)
                    if qty_match_full:
                        qty_str = qty_match_full.group(1).replace('.', '')
                        if current_code in product_code_to_vasilhame_map:
                            parsed_data.append({'PRODUTO_CODE': current_code, 'QUANTIDADE': int(qty_str)})
                        current_code = None
                    else:
                         qty_match_end = re.search(r'\s+([\d\.]+)$', line)
                         if qty_match_end:
                             qty_str = qty_match_end.group(1).replace('.', '')
                             if current_code in product_code_to_vasilhame_map:
                                 parsed_data.append({'PRODUTO_CODE': current_code, 'QUANTIDADE': int(qty_str)})
                             current_code = None
                elif current_code:
                    qty_match_next = re.search(r'([\d\.]+)\s+[\d\.]+,\d+', line)
                    if qty_match_next:
                        qty_str = qty_match_next.group(1).replace('.', '')
                        if current_code in product_code_to_vasilhame_map:
                            parsed_data.append({'PRODUTO_CODE': current_code, 'QUANTIDADE': int(qty_str)})
                        current_code = None 
            if not parsed_data: return None, effective_date_str, effective_date_full
            df_estoque = pd.DataFrame(parsed_data)
            df_estoque['Vasilhame'] = df_estoque['PRODUTO_CODE'].map(product_code_to_vasilhame_map)
            df_txt_qty = df_estoque.groupby('Vasilhame')['QUANTIDADE'].sum().reset_index()
            df_txt_qty.rename(columns={'QUANTIDADE': 'Qtd_emprestimo'}, inplace=True)
            return df_txt_qty, effective_date_str, effective_date_full

        def process_pdf_content(pdf_file, product_map):
            parsed_data = []
            filename_match = re.search(r'([a-zA-Z\s]+)\s+(\d{2}-\d{2}-\d{4})\.pdf', pdf_file.name)
            if not filename_match: st.error(f"Erro no nome do arquivo PDF: {pdf_file.name}"); return pd.DataFrame()
            source_name = filename_match.group(1).strip()
            date_str = filename_match.group(2) 
            effective_date_obj = datetime.strptime(date_str, '%d-%m-%Y')
            effective_date_str = effective_date_obj.strftime('%d/%m')
            effective_date_full = effective_date_obj.date()
            source_to_col_map = {'PONTA GROSSA': 'Ponta Grossa (0328)', 'ARARAQUARA': 'Araraquara (0336)', 'ITU': 'Itu (0002)'}
            col_suffix = source_to_col_map.get(source_name.upper(), source_name)
            pdf_reader = PyPDF2.PdfReader(io.BytesIO(pdf_file.getvalue()))
            pdf_content = ""
            for page in pdf_reader.pages: pdf_content += page.extract_text()
            data_line_pattern = re.compile(r'^\s*"?(\d{15,})[^"\n]*?"?.*?"?([-+]?[\d.,]+)"?\s*$', re.MULTILINE)
            for line_match in data_line_pattern.finditer(pdf_content):
                material_code = line_match.group(1).strip()
                saldo_str = line_match.group(2).replace('.', '').replace(',', '.').strip()
                try: saldo = float(saldo_str)
                except ValueError: saldo = 0.0
                if material_code in product_map:
                    vasilhame = product_map[material_code]
                    credito = abs(saldo) if saldo < 0 else 0.0
                    debito = saldo if saldo >= 0 else 0.0
                    parsed_data.append({'Vasilhame': vasilhame, 'Dia': effective_date_str, 'DataCompleta': effective_date_full, f'Credito {col_suffix}': credito, f'Debito {col_suffix}': debito})
            if not parsed_data: st.warning(f"Nenhum dado encontrado no PDF: {pdf_file.name}"); return pd.DataFrame()
            df_parsed = pd.DataFrame(parsed_data)
            pdf_value_cols = [col for col in df_parsed.columns if 'Credito' in col or 'Debito' in col]
            agg_dict = {col: 'sum' for col in pdf_value_cols}; agg_dict['DataCompleta'] = 'max'
            return df_parsed.groupby(['Vasilhame', 'Dia'], as_index=False).agg(agg_dict)
        
        uploaded_txt_files = st.file_uploader("Envie os arquivos TXT de empréstimos (Ex: ESTOQUE0102.TXT)", type=["txt"], accept_multiple_files=True, key="vasil_txt_uploader") 
        uploaded_vendas_files = st.file_uploader("Envie os arquivos TXT de Vendas (Ex: VENDA2411.TXT) [Opcional]", type=["txt"], accept_multiple_files=True, key="vasil_vendas_uploader")
        uploaded_excel_contagem = st.file_uploader("Envie o arquivo Excel de contagem (Ex: Contagem Vasilhames.xlsx)", type=["xlsx"], key="vasil_excel_uploader")
        uploaded_pdf_files = st.file_uploader("Envie os arquivos PDF de fábrica (Ex: PONTA GROSSA 07-11-2025.pdf)", type=["pdf"], accept_multiple_files=True, key="vasil_pdf_uploader")
        
        if st.button("Processar e Consolidar Dados"):
            if uploaded_txt_files and uploaded_excel_contagem is not None:
                try:
                    st.info("Sincronizando com Google Sheets e processando arquivos...")
                    
                    try:
                        df_old_txt_data = load_from_gsheets(sheet_client, 'txt_data')
                        df_old_pdf_data = load_from_gsheets(sheet_client, 'pdf_data')
                        df_old_vendas_data = load_from_gsheets(sheet_client, 'vendas_data')
                        df_old_excel_data = load_from_gsheets(sheet_client, 'excel_data')
                    except Exception as e:
                        st.error(f"Erro ao baixar dados da nuvem. Tente limpar e reiniciar. Detalhe: {e}")
                        st.stop()

                    def sanear_dataframe(df, col_valor=None):
                        if df.empty: return df
                        if 'DataCompleta' in df.columns:
                            df['DataCompleta'] = pd.to_datetime(df['DataCompleta'], errors='coerce')
                            df = df.dropna(subset=['DataCompleta'])
                            df['Dia'] = df['DataCompleta'].dt.strftime('%d/%m')
                        elif 'Dia' in df.columns:
                            try:
                                temp_date = pd.to_datetime(df['Dia'], errors='coerce')
                                df.loc[temp_date.notna(), 'Dia'] = temp_date.dt.strftime('%d/%m')
                            except: pass

                        if col_valor and col_valor in df.columns:
                            if df[col_valor].dtype == object:
                                df[col_valor] = df[col_valor].astype(str).str.replace(',', '.')
                            df[col_valor] = pd.to_numeric(df[col_valor], errors='coerce').fillna(0)
                        return df

                    df_old_txt_data = sanear_dataframe(df_old_txt_data, col_valor='Qtd_emprestimo')
                    df_old_pdf_data = sanear_dataframe(df_old_pdf_data) 
                    df_old_vendas_data = sanear_dataframe(df_old_vendas_data, col_valor='Vendas')
                    
                    if not df_old_excel_data.empty:
                        col_data = 'DataCompleta'
                        if 'DataCompleta_excel' in df_old_excel_data.columns: col_data = 'DataCompleta_excel'
                        if col_data in df_old_excel_data.columns:
                            df_old_excel_data[col_data] = pd.to_datetime(df_old_excel_data[col_data], errors='coerce')
                            df_old_excel_data.dropna(subset=[col_data], inplace=True)
                            df_old_excel_data['Dia'] = df_old_excel_data[col_data].dt.strftime('%d/%m')

                    if not df_old_excel_data.empty:
                        if 'DataCompleta' in df_old_excel_data.columns and 'DataCompleta_excel' not in df_old_excel_data.columns:
                             df_old_excel_data.rename(columns={'DataCompleta': 'DataCompleta_excel'}, inplace=True)

                    new_txt_data_list = []
                    for uploaded_txt_file in uploaded_txt_files:
                        df_txt_qty, effective_date_str, effective_date_full = process_txt_file_st(uploaded_txt_file)
                        if df_txt_qty is not None:
                            df_txt_qty['Dia'] = effective_date_str
                            df_txt_qty['DataCompleta'] = effective_date_full
                            new_txt_data_list.append(df_txt_qty)
                    
                    if new_txt_data_list:
                        df_new_txt = pd.concat(new_txt_data_list, ignore_index=True)
                        df_all_txt_combined = pd.concat([df_old_txt_data, df_new_txt], ignore_index=True)
                        if 'DataCompleta' in df_all_txt_combined.columns: 
                            df_all_txt_combined['DataCompleta'] = pd.to_datetime(df_all_txt_combined['DataCompleta'], errors='coerce')
                        
                        df_all_processed_txt_data = df_all_txt_combined.groupby(['Vasilhame', 'Dia']).agg(
                            Qtd_emprestimo=('Qtd_emprestimo', 'sum'), 
                            DataCompleta=('DataCompleta', 'max')
                        ).reset_index()
                        
                        save_to_gsheets(sheet_client, 'txt_data', df_all_processed_txt_data)
                        st.success("TXT: Dados atualizados na Nuvem!")
                    else: 
                        df_all_processed_txt_data = df_old_txt_data 
                    
                    new_vendas_data_list = []
                    if uploaded_vendas_files:
                        for v_file in uploaded_vendas_files:
                            df_v = process_vendas_file(v_file)
                            if df_v is not None: new_vendas_data_list.append(df_v)
                    
                    if new_vendas_data_list:
                        df_new_vendas = pd.concat(new_vendas_data_list, ignore_index=True)
                        df_all_vendas_combined = pd.concat([df_old_vendas_data, df_new_vendas], ignore_index=True)
                        if 'DataCompleta' in df_all_vendas_combined.columns: 
                            df_all_vendas_combined['DataCompleta'] = pd.to_datetime(df_all_vendas_combined['DataCompleta'], errors='coerce')
                        
                        df_all_processed_vendas_data = df_all_vendas_combined.groupby(['Vasilhame', 'Dia']).agg(
                            Vendas=('Vendas', 'sum'), 
                            DataCompleta=('DataCompleta', 'max')
                        ).reset_index()
                        
                        save_to_gsheets(sheet_client, 'vendas_data', df_all_processed_vendas_data)
                        st.success("Vendas: Dados atualizados na Nuvem!")
                    else:
                        df_all_processed_vendas_data = df_old_vendas_data
                    
                    if df_all_processed_vendas_data.empty:
                         df_all_processed_vendas_data = pd.DataFrame(columns=['Vasilhame', 'Dia', 'Vendas', 'DataCompleta'])

                    new_pdf_data_list = []
                    if uploaded_pdf_files:
                        pdf_map = {
                            '000000000000215442': '587-002 - CAIXA PLASTICA HEINEKEN 600ML', 
                            '000000000000215208': '587-002 - CAIXA PLASTICA HEINEKEN 600ML', 
                            '000000000000381411': '591-002 - CAIXA PLASTICA HEINEKEN 330ML', 
                            '000000000000107380': '555-001 - CAIXA PLASTICA 1L', 
                            '000000000000152598': '546-004 - CAIXA PLASTICA 24UN 300ML', 
                            '000000000000000470': '550-001 - CAIXA PLASTICA 600ML',
                            '000000000000048261': '563-008 - BARRIL INOX 30L', 
                            '000000000000048272': '564-009 - BARRIL INOX 50L',
                            '000000000000185039': CRATE_TO_BOTTLE_MAP['546-004 - CAIXA PLASTICA 24UN 300ML'],
                            '000000000000002496': NAME_540_001, 
                            '000000000000107523': CRATE_TO_BOTTLE_MAP['555-001 - CAIXA PLASTICA 1L'],
                            '000000000000152592': CRATE_TO_BOTTLE_MAP['546-004 - CAIXA PLASTICA 24UN 300ML'],
                            '000000000000215443': CRATE_TO_BOTTLE_MAP['587-002 - CAIXA PLASTICA HEINEKEN 600ML'],
                            '000000000000381408': CRATE_TO_BOTTLE_MAP['591-002 - CAIXA PLASTICA HEINEKEN 330ML'],
                            '000000000000152597': CRATE_TO_BOTTLE_MAP['546-004 - CAIXA PLASTICA 24UN 300ML'], 
                            '000000000000000471': NAME_540_001,      
                            '000000000000107522': CRATE_TO_BOTTLE_MAP['555-001 - CAIXA PLASTICA 1L'],        
                            '000000000000215209': CRATE_TO_BOTTLE_MAP['587-002 - CAIXA PLASTICA HEINEKEN 600ML'], 
                            '000000000000381409': CRATE_TO_BOTTLE_MAP['591-002 - CAIXA PLASTICA HEINEKEN 330ML']  
                        }
                        for pdf_file in uploaded_pdf_files:
                            df_pdf_current = process_pdf_content(pdf_file, pdf_map)
                            if not df_pdf_current.empty: new_pdf_data_list.append(df_pdf_current)
                    
                    if new_pdf_data_list:
                        df_new_pdf = pd.concat(new_pdf_data_list, ignore_index=True)
                        df_all_pdf_combined = pd.concat([df_old_pdf_data, df_new_pdf], ignore_index=True)
                        pdf_value_cols = [col for col in df_all_pdf_combined.columns if 'Credito' in col or 'Debito' in col]
                        df_all_pdf_combined[pdf_value_cols] = df_all_pdf_combined[pdf_value_cols].fillna(0)
                        if 'DataCompleta' in df_all_pdf_combined.columns: 
                            df_all_pdf_combined['DataCompleta'] = pd.to_datetime(df_all_pdf_combined['DataCompleta'], errors='coerce')
                        
                        agg_dict = {col: 'sum' for col in pdf_value_cols}; agg_dict['DataCompleta'] = 'max' 
                        if pdf_value_cols: 
                            df_all_processed_pdf_data = df_all_pdf_combined.groupby(['Vasilhame', 'Dia'], as_index=False).agg(agg_dict)
                        else: 
                            df_all_processed_pdf_data = df_all_pdf_combined.groupby(['Vasilhame', 'Dia'], as_index=False).agg(DataCompleta=('DataCompleta', 'max')).reset_index()
                        
                        save_to_gsheets(sheet_client, 'pdf_data', df_all_processed_pdf_data)
                        st.success("PDF: Dados atualizados na Nuvem!")
                    else: 
                        df_all_processed_pdf_data = df_old_pdf_data
                    
                    if df_all_processed_txt_data.empty: df_all_processed_txt_data = pd.DataFrame(columns=['Vasilhame', 'Dia', 'Qtd_emprestimo', 'DataCompleta'])
                    if df_all_processed_pdf_data.empty: df_all_processed_pdf_data = pd.DataFrame(columns=['Vasilhame', 'Dia', 'DataCompleta'])

                    df_contagem = pd.read_excel(uploaded_excel_contagem, sheet_name='Respostas ao formulário 1')
                    df_contagem['Carimbo de data/hora'] = pd.to_datetime(df_contagem['Carimbo de data/hora'])
                    df_contagem['DataCompleta'] = df_contagem['Carimbo de data/hora'].dt.date
                    df_contagem['Dia'] = df_contagem['Carimbo de data/hora'].dt.strftime('%d/%m')
                    
                    def map_excel_names_and_get_target(name):
                        name_upper = str(name).upper()
                        target_crate = name 
                        target_bottle = None
                        factor = 1
                        if '063-005' in name_upper: target_bottle = '546-001 - GARRAFA 300ML'; return None, target_bottle, 1
                        if '540-001' in name_upper: target_bottle = NAME_540_001; return None, target_bottle, 1
                        if '541-002' in name_upper: target_bottle = '541-002 - GARRAFA 1L'; return None, target_bottle, 1
                        if '586-001' in name_upper: target_bottle = '586-001 - GARRAFA HEINEKEN 600ML'; return None, target_bottle, 1
                        if '593-001' in name_upper: target_bottle = '593-001 - GARRAFA HEINEKEN 330ML'; return None, target_bottle, 1

                        if '550-012' in name_upper or 'EISENBAHN' in name_upper or '550-001' in name_upper or 'MISTA' in name_upper or 'AMBEV' in name_upper or 'CINZA' in name_upper:
                             target_crate = '550-001 - CAIXA PLASTICA 600ML'
                        elif '587-002' in name_upper or ('HEINEKEN' in name_upper and '600' in name_upper): target_crate = '587-002 - CAIXA PLASTICA HEINEKEN 600ML'
                        elif '546-004' in name_upper: target_crate = '546-004 - CAIXA PLASTICA 24UN 300ML'
                        elif '591-002' in name_upper: target_crate = '591-002 - CAIXA PLASTICA HEINEKEN 330ML'
                        elif '555-001' in name_upper: target_crate = '555-001 - CAIXA PLASTICA 1L'

                        if target_crate in CRATE_TO_BOTTLE_MAP:
                            target_bottle = CRATE_TO_BOTTLE_MAP[target_crate]
                            factor = FACTORS.get(target_crate, 1)
                            
                        return target_crate, target_bottle, factor

                    def calculate_assets(row):
                        target_crate, target_bottle, factor = map_excel_names_and_get_target(row['Qual vasilhame ?'])
                        garrafa_cheia = 0.0; caixa_vazia = 0.0; caixa_cheia = 0.0
                        
                        if 'Quantidade estoque cheias?' in row.index:
                             def get_val(col):
                                 try: return float(row.get(col, 0) or 0)
                                 except: return 0.0
                             
                             qtd_cheias = get_val('Quantidade estoque cheias?')
                             qtd_vazias = get_val('Quantidade estoque vazias?')
                             transito_cheias = get_val('Em transito cheias (Entrega)?')
                             transito_vazias = get_val('Em transito vazias (Entrega)?')
                             carreta = get_val('Em transito (carreta)?')
                             
                             total_cheias = qtd_cheias + transito_cheias + carreta
                             total_vazias = qtd_vazias + transito_vazias
                             
                             if target_crate is None and target_bottle is not None:
                                 garrafa_cheia = total_cheias + total_vazias
                                 caixa_cheia = 0
                                 caixa_vazia = 0
                             elif target_bottle:
                                 garrafa_cheia = total_cheias * factor
                                 caixa_cheia = total_cheias
                                 caixa_vazia = total_vazias
                             else:
                                 caixa_cheia = total_cheias
                                 caixa_vazia = total_vazias
                        else:
                            if 'Total' in row.index and pd.notnull(row['Total']): total = float(row['Total'])
                            else: total = float(row.get('Quantidade estoque ?', 0) or 0) + float(row.get('Em transito (Entrega)?', 0) or 0) + float(row.get('Em transito (carreta)?', 0) or 0)
                            
                            if target_crate is None and target_bottle is not None: garrafa_cheia = total
                            elif target_bottle: garrafa_cheia = total * factor; caixa_cheia = total
                            else: caixa_cheia = total

                        return pd.Series([target_crate, target_bottle, garrafa_cheia, caixa_vazia, caixa_cheia], index=['TargetCrate', 'TargetBottle', 'GarrafaCheia', 'CaixaVazia', 'CaixaCheia'])

                    df_contagem[['TargetCrate', 'TargetBottle', 'GarrafaCheia', 'CaixaVazia', 'CaixaCheia']] = df_contagem.apply(calculate_assets, axis=1)

                    def calculate_assets_converted(row):
                        target_crate, target_bottle, factor = map_excel_names_and_get_target(row['Qual vasilhame ?'])
                        
                        qc = float(row.get('Quantidade estoque cheias?', 0) or 0)
                        qv = float(row.get('Quantidade estoque vazias?', 0) or 0)
                        tc = float(row.get('Em transito cheias (Entrega)?', 0) or 0)
                        tv = float(row.get('Em transito vazias (Entrega)?', 0) or 0)
                        car = float(row.get('Em transito (carreta)?', 0) or 0)
                        
                        g_qc, g_qv, g_tc, g_tv, g_car = 0,0,0,0,0
                        c_qc, c_qv, c_tc, c_tv, c_car = 0,0,0,0,0
                        
                        if target_crate is None and target_bottle is not None:
                            g_qc = qc; g_qv = qv; g_tc = tc; g_tv = tv; g_car = car
                        elif target_bottle:
                            g_qc = qc * factor
                            g_tc = tc * factor
                            g_car = car * factor
                            c_qc = qc; c_qv = qv; c_tc = tc; c_tv = tv; c_car = car
                        else:
                            c_qc = qc; c_qv = qv; c_tc = tc; c_tv = tv; c_car = car
                        
                        return pd.Series([target_crate, target_bottle, g_qc, g_qv, g_tc, g_tv, g_car, c_qc, c_qv, c_tc, c_tv, c_car], 
                                         index=['TargetCrate', 'TargetBottle', 'G_QC', 'G_QV', 'G_TC', 'G_TV', 'G_CAR', 'C_QC', 'C_QV', 'C_TC', 'C_TV', 'C_CAR'])

                    df_contagem[['TargetCrate', 'TargetBottle', 'G_QC', 'G_QV', 'G_TC', 'G_TV', 'G_CAR', 'C_QC', 'C_QV', 'C_TC', 'C_TV', 'C_CAR']] = df_contagem.apply(calculate_assets_converted, axis=1)

                    agg_cols_g = {'G_QC':'sum', 'G_QV':'sum', 'G_TC':'sum', 'G_TV':'sum', 'G_CAR':'sum', 'Carimbo de data/hora':'max'}
                    df_agg_garrafa = df_contagem.dropna(subset=['TargetBottle']).groupby(['TargetBottle', 'Dia']).agg(agg_cols_g).reset_index()
                    df_agg_garrafa.rename(columns={'TargetBottle': 'Vasilhame', 'G_QC':'Quantidade estoque cheias', 'G_QV':'Quantidade estoque vazias', 'G_TC':'Em transito cheias (Entrega)', 'G_TV':'Em transito vazias (Entrega)', 'G_CAR':'Em transito (carreta)'}, inplace=True)
                    df_agg_garrafa['Contagem Cheias'] = df_agg_garrafa['Quantidade estoque cheias'] + df_agg_garrafa['Em transito cheias (Entrega)'] + df_agg_garrafa['Em transito (carreta)']
                    df_agg_garrafa['Contagem Vazias'] = df_agg_garrafa['Quantidade estoque vazias'] + df_agg_garrafa['Em transito vazias (Entrega)']

                    agg_cols_c = {'C_QC':'sum', 'C_QV':'sum', 'C_TC':'sum', 'C_TV':'sum', 'C_CAR':'sum', 'Carimbo de data/hora':'max'}
                    df_agg_caixa = df_contagem.dropna(subset=['TargetCrate']).groupby(['TargetCrate', 'Dia']).agg(agg_cols_c).reset_index()
                    df_agg_caixa.rename(columns={'TargetCrate': 'Vasilhame', 'C_QC':'Quantidade estoque cheias', 'C_QV':'Quantidade estoque vazias', 'C_TC':'Em transito cheias (Entrega)', 'C_TV':'Em transito vazias (Entrega)', 'C_CAR':'Em transito (carreta)'}, inplace=True)
                    df_agg_caixa['Contagem Cheias'] = df_agg_caixa['Quantidade estoque cheias'] + df_agg_caixa['Em transito cheias (Entrega)'] + df_agg_caixa['Em transito (carreta)']
                    df_agg_caixa['Contagem Vazias'] = df_agg_caixa['Quantidade estoque vazias'] + df_agg_caixa['Em transito vazias (Entrega)']

                    df_excel_agg = pd.concat([df_agg_garrafa, df_agg_caixa], ignore_index=True)
                    df_excel_agg.rename(columns={'DataCompleta': 'DataCompleta_excel'}, inplace=True)

                    if not df_old_excel_data.empty:
                         for col in df_excel_agg.columns:
                             if col not in df_old_excel_data.columns: df_old_excel_data[col] = 0
                         if 'DataCompleta_excel' in df_old_excel_data.columns: 
                             df_old_excel_data['DataCompleta_excel'] = pd.to_datetime(df_old_excel_data['DataCompleta_excel'], errors='coerce')
                         
                         df_excel_agg = pd.concat([df_old_excel_data, df_excel_agg]).drop_duplicates(subset=['Vasilhame', 'Dia'], keep='last').reset_index(drop=True)
                    
                    save_to_gsheets(sheet_client, 'excel_data', df_excel_agg)
                    st.success("Contagem Excel: Dados atualizados na Nuvem!")

                    def forcar_formato_visual(df):
                        if df.empty: return df
                        df = df.copy()
                        
                        if 'DataCompleta' in df.columns:
                            df['DataCompleta'] = pd.to_datetime(df['DataCompleta'], errors='coerce')
                            mask_valid = df['DataCompleta'].notna()
                            if mask_valid.any():
                                df.loc[mask_valid, 'Dia'] = df.loc[mask_valid, 'DataCompleta'].dt.strftime('%d/%m')
                        
                        try:
                            temp_dates = pd.to_datetime(df['Dia'], format='%d/%m', errors='coerce')
                            mask_nat = temp_dates.isna()
                            if mask_nat.any():
                                temp_dates.loc[mask_nat] = pd.to_datetime(df.loc[mask_nat, 'Dia'], errors='coerce')
                            df.loc[temp_dates.notna(), 'Dia'] = temp_dates.dt.strftime('%d/%m')
                        except: pass
                        return df

                    df_excel_agg = forcar_formato_visual(df_excel_agg)
                    df_all_processed_txt_data = forcar_formato_visual(df_all_processed_txt_data)
                    df_all_processed_pdf_data = forcar_formato_visual(df_all_processed_pdf_data)
                    df_all_processed_vendas_data = forcar_formato_visual(df_all_processed_vendas_data)

                    all_dates = set()
                    if not df_excel_agg.empty: all_dates.update(df_excel_agg['Dia'].dropna().unique())
                    if not df_all_processed_txt_data.empty: all_dates.update(df_all_processed_txt_data['Dia'].dropna().unique())
                    if not df_all_processed_pdf_data.empty: all_dates.update(df_all_processed_pdf_data['Dia'].dropna().unique())
                    if not all_dates: all_dates.add(datetime.now().strftime('%d/%m'))

                    required_vasilhames = set(list(FACTORS.keys()) + list(CRATE_TO_BOTTLE_MAP.values()))
                    
                    if not df_all_processed_txt_data.empty and 'Vasilhame' in df_all_processed_txt_data.columns:
                        required_vasilhames.update(df_all_processed_txt_data['Vasilhame'].dropna().unique())
                    
                    if not df_all_processed_pdf_data.empty and 'Vasilhame' in df_all_processed_pdf_data.columns:
                        required_vasilhames.update(df_all_processed_pdf_data['Vasilhame'].dropna().unique())
                        
                    if not df_excel_agg.empty and 'Vasilhame' in df_excel_agg.columns:
                        required_vasilhames.update(df_excel_agg['Vasilhame'].dropna().unique())

                    skeleton_rows = []
                    sorted_products = sorted(list(required_vasilhames)) 
                    
                    for prod in sorted_products:
                        for day in sorted(list(all_dates)):
                             skeleton_rows.append({'Vasilhame': prod, 'Dia': day})
                    df_skeleton = pd.DataFrame(skeleton_rows)

                    df_final = df_skeleton.copy()
                    
                    df_final = pd.merge(df_final, df_excel_agg, on=['Vasilhame', 'Dia'], how='left')
                    df_final = pd.merge(df_final, df_all_processed_txt_data, on=['Vasilhame', 'Dia'], how='left')
                    df_final = pd.merge(df_final, df_all_processed_pdf_data, on=['Vasilhame', 'Dia'], how='left')
                    df_final = pd.merge(df_final, df_all_processed_vendas_data, on=['Vasilhame', 'Dia'], how='left')
                    
                    cols_data = [c for c in df_final.columns if 'DataCompleta' in c]
                    df_final['DataCompleta'] = pd.NaT
                    for col in cols_data:
                        df_final['DataCompleta'] = df_final['DataCompleta'].fillna(pd.to_datetime(df_final[col], errors='coerce'))
                        if col != 'DataCompleta': df_final.drop(col, axis=1, inplace=True)

                    def infer_date(row):
                        if pd.isna(row['DataCompleta']):
                            try: 
                                current_year = datetime.now().year
                                return datetime.strptime(f"{row['Dia']}/{current_year}", "%d/%m/%Y")
                            except: return pd.NaT
                        return row['DataCompleta']
                    
                    df_final['DataCompleta'] = df_final.apply(infer_date, axis=1)

                    output_form_cols = ['Quantidade estoque cheias', 'Quantidade estoque vazias', 'Em transito cheias (Entrega)', 'Em transito vazias (Entrega)', 'Em transito (carreta)']
                    numeric_cols = ['Contagem Cheias', 'Contagem Vazias', 'Qtd_emprestimo', 'Vendas'] + output_form_cols + [col for col in df_final.columns if 'Credito' in col or 'Debito' in col]
                    
                    for col in numeric_cols:
                        if col in df_final.columns: 
                            df_final[col] = pd.to_numeric(df_final[col], errors='coerce').fillna(0)
                        else:
                            df_final[col] = 0

                    if 'Vendas' not in df_final.columns: df_final['Vendas'] = 0

                    groupby_cols = ['Vasilhame', 'Dia', 'DataCompleta']
                    cols_to_sum = [c for c in numeric_cols if c in df_final.columns]
                    df_final = df_final.groupby(groupby_cols)[cols_to_sum].sum().reset_index()

                    df_final['Total Revenda'] = df_final['Qtd_emprestimo'] + df_final['Contagem Cheias'] + df_final['Contagem Vazias'] + df_final.filter(like='Credito').sum(axis=1) - df_final.filter(like='Debito').sum(axis=1) + df_final['Vendas']
                    
                    df_final.sort_values(by=['Vasilhame', 'DataCompleta'], inplace=True, na_position='first')
                    
                    def calcular_diferenca_regra_negocio(grupo):
                        data_base_travamento = pd.to_datetime('2025-11-05')
                        data_inicio_calculo = pd.to_datetime('2025-11-10')
                        mask_base = grupo['DataCompleta'] >= data_base_travamento
                        dados_base = grupo.loc[mask_base]
                        if not dados_base.empty: estoque_travado = dados_base.iloc[0]['Total Revenda']
                        else: estoque_travado = 0
                        diferencas = pd.Series(0.0, index=grupo.index)
                        mask_calculo = grupo['DataCompleta'] >= data_inicio_calculo
                        if estoque_travado != 0: 
                            diferencas[mask_calculo] = grupo.loc[mask_calculo, 'Total Revenda'] - estoque_travado
                        grupo['Diferença'] = diferencas
                        return grupo

                    df_final = df_final.groupby('Vasilhame', group_keys=False).apply(calcular_diferenca_regra_negocio)
                    df_final_output = df_final.drop('DataCompleta', axis=1)

                    output_cols = [c for c in df_final_output.columns if c not in ['Diferença', 'Vendas']]
                    df_final_output = df_final_output[output_cols + ['Diferença', 'Vendas']]
                    
                    st.subheader("✅ Tabela Consolidada de Vasilhames")
                    st.dataframe(df_final_output)
                    
                    st.info("Salvando tabela consolidada na Nuvem...")
                    save_to_gsheets(sheet_client, 'CONSOLIDADO_GERAL', df_final_output)
                    st.success("Tabela Consolidada salva na aba 'CONSOLIDADO_GERAL'!")

                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                        df_final_output.to_excel(writer, sheet_name='GERAL', index=False)
                        unique_products = df_final_output['Vasilhame'].unique()
                        
                        caixas_list = sorted([p for p in unique_products if 'CAIXA' in str(p).upper() or 'BARRIL' in str(p).upper() or 'CILINDRO' in str(p).upper()])
                        garrafas_list = sorted([p for p in unique_products if 'GARRAFA' in str(p).upper()])
                        outros_list = sorted([p for p in unique_products if p not in caixas_list and p not in garrafas_list])
                        sorted_products = caixas_list + garrafas_list + outros_list

                        for product in sorted_products:
                            df_product = df_final_output[df_final_output['Vasilhame'] == product]
                            safe_sheet_name = str(product).replace('/', '-').replace('\\', '-').replace('?', '').replace('*', '').replace('[', '').replace(']', '').replace(':', '')[:31]
                            df_product.to_excel(writer, sheet_name=safe_sheet_name, index=False)
                            
                    output.seek(0)
                    st.download_button(label="📥 Baixar Tabela Consolidada", data=output, file_name="Vasilhames_Consolidado.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

                except Exception as e:
                    st.error(f"Ocorreu um erro durante o processamento: {e}")

    # --- SCRIPT ABASTECIMENTO ---
    elif script_choice == "Abastecimento":
        st.subheader("Análise de Abastecimento")
        uploaded_file = st.file_uploader("Envie o arquivo de abastecimento (.xlsx ou .csv)", type=["xlsx", "csv"], key="abastec_uploader") 
        if uploaded_file is not None:
            try:
                st.info("Processando...")
                try:
                    if uploaded_file.name.endswith('.csv'): df = pd.read_csv(uploaded_file)
                    elif uploaded_file.name.endswith('.xlsx'): df = pd.read_excel(uploaded_file)
                    else: st.error("Formato não suportado."); return
                except Exception as e: st.error(f"Erro ao ler arquivo: {e}"); return

                df.columns = [col.upper().strip().replace('HORA', 'HORÁRIO') for col in df.columns]
                column_mapping = {
                    'DATA ABASTECIMENTO': ['DATA', 'DATA ABASTECIMENTO', 'DATE', 'DATA_ABASTECIMENTO'],
                    'HORÁRIO': ['HORÁRIO', 'HORA', 'HORA DO ABASTECIMENTO'],
                    'TIPO DE ABASTECIMENTO': ['TIPO DE ABASTECIMENTO', 'TIPO_ABASTECIMENTO', 'COMBUSTÍVEL', 'TIPO'],
                    'PLACA': ['PLACA', 'PLACA_VEICULO'],
                    'KM': ['KM', 'QUILOMETRAGEM'],
                    'LITROS': ['LITROS', 'VOLUME'],
                    'MOTORISTA': ['MOTORISTA', 'RESPONSÁVEL'],
                }
                df_unified = pd.DataFrame()
                for new_name, possible_names in column_mapping.items():
                    for old_name in possible_names:
                        if old_name.upper() in df.columns: df_unified[new_name] = df[old_name.upper()]; break
                    else: st.warning(f"Coluna '{new_name}' não encontrada."); df_unified[new_name] = np.nan
                df = df_unified
                df['DATA ABASTECIMENTO'] = pd.to_datetime(df['DATA ABASTECIMENTO'], errors='coerce').dt.date
                df['HORÁRIO'] = pd.to_datetime(df['HORÁRIO'], format='%H:%M:%S', errors='coerce').dt.time
                df['KM'] = pd.to_numeric(df['KM'], errors='coerce')
                df['LITROS'] = pd.to_numeric(df['LITROS'], errors='coerce')
                df.dropna(subset=['DATA ABASTECIMENTO', 'KM', 'LITROS'], inplace=True)
                
                df_diesel = df[df['TIPO DE ABASTECIMENTO'].str.upper() == 'DIESEL'].copy()
                if not df_diesel.empty:
                    excel_data_diesel = io.BytesIO()
                    with pd.ExcelWriter(excel_data_diesel, engine='xlsxwriter') as writer:
                        placas_diesel = sorted(df_diesel['PLACA'].unique())
                        for placa in placas_diesel:
                            df_placa = df_diesel[df_diesel['PLACA'] == placa].copy()
                            df_placa.sort_values(by=['DATA ABASTECIMENTO', 'HORÁRIO'], ascending=True, inplace=True)
                            df_placa['DISTANCIA_PERCORRIDA'] = df_placa['KM'].diff()
                            df_placa['MEDIA_LITROS_KM'] = df_placa['LITROS'] / df_placa['DISTANCIA_PERCORRIDA']
                            df_placa['ALERTA KM'] = ''
                            df_placa.loc[df_placa['DISTANCIA_PERCORRIDA'] < 0, 'ALERTA KM'] = 'ALERTA: KM menor'
                            df_placa_output = df_placa.rename(columns={'DATA ABASTECIMENTO': 'Data Abastecimento', 'MEDIA_LITROS_KM': 'Média de litros por KM'})
                            df_placa_output.to_excel(writer, sheet_name=placa, index=False)
                    excel_data_diesel.seek(0)
                    st.success("Planilha de Diesel OK!")
                    st.download_button(label="📥 Baixar Diesel", data=excel_data_diesel, file_name="planilha_diesel.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                
                df_arla = df[df['TIPO DE ABASTECIMENTO'].str.upper() == 'ARLA'].copy()
                if not df_arla.empty:
                    excel_data_arla = io.BytesIO()
                    with pd.ExcelWriter(excel_data_arla, engine='xlsxwriter') as writer:
                        placas_arla = sorted(df_arla['PLACA'].unique())
                        for placa in placas_arla:
                            df_placa = df_arla[df_arla['PLACA'] == placa].copy()
                            df_placa.sort_values(by=['DATA ABASTECIMENTO', 'HORÁRIO'], ascending=True, inplace=True)
                            df_placa['DISTANCIA_PERCORRIDA'] = df_placa['KM'].diff()
                            df_placa['MEDIA_LITROS_KM'] = df_placa['LITROS'] / df_placa['DISTANCIA_PERCORRIDA']
                            df_placa['ALERTA KM'] = ''
                            df_placa.loc[df_placa['DISTANCIA_PERCORRIDA'] < 0, 'ALERTA KM'] = 'ALERTA: KM menor'
                            df_placa_output = df_placa.rename(columns={'DATA ABASTECIMENTO': 'Data Abastecimento', 'MEDIA_LITROS_KM': 'Média de litros por KM'})
                            df_placa_output.to_excel(writer, sheet_name=placa, index=False)
                    excel_data_arla.seek(0)
                    st.success("Planilha de Arla OK!")
                    st.download_button(label="📥 Baixar Arla", data=excel_data_arla, file_name="planilha_arla.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            except Exception as e:
                st.error(f"Ocorreu um erro ao processar o arquivo: {e}")

    # --- SCRIPT MANUTENÇÃO VEÍCULOS (NOVA LÓGICA ROBUSTA COM FILTRO DE PÁGINA) ---
    elif script_choice == "Manutenção Veículos":
        st.subheader("Manutenção de Veículos (FleetCom)")
        st.info("Envie os relatórios em PDF. O sistema irá consolidar tudo em um único arquivo Excel com abas separadas.")

        col1, col2, col3 = st.columns(3)
        with col1:
            up_cam = st.file_uploader("🚛 Caminhões (PDF)", type=["pdf"], key="up_cam")
        with col2:
            up_car = st.file_uploader("🚗 Carros (PDF)", type=["pdf"], key="up_car")
        with col3:
            up_mot = st.file_uploader("🏍️ Motos (PDF)", type=["pdf"], key="up_mot")

        def extrair_dados_fleetcom(pdf_buffer):
            pdf_reader = PyPDF2.PdfReader(io.BytesIO(pdf_buffer.getvalue()))
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"

            # Limpeza profunda
            text = text.replace('\r\n', '\n').replace('"\n","', ' ').replace('",\n"', ' ')
            text = text.replace('"\n"', '\n').replace('"', '')
            text = re.sub(r'^\s*,\s*', '', text, flags=re.MULTILINE)

            marcadores_resumo = ["N. DE VEÍCULOS ATENDIDOS", "NO. DE IM'S PREVENTIVAS", "RESUMO POR PLACA"]
            for marcador in marcadores_resumo:
                idx_resumo = text.upper().find(marcador)
                if idx_resumo != -1:
                    text = text[:idx_resumo]

            parsed_data = []

            # Passo 1: Encontra com segurança onde cada veículo começa
            veiculo_pattern = re.compile(r'([A-Z]{3}-?\d[A-Z0-9]\d{2})\s+(.+?)(?:\s+[A-Z]{3}-?\d[A-Z0-9]\d{2})?\s+([\d\.]+,\d+)\s+(\d{4})')
            matches_veiculos = list(veiculo_pattern.finditer(text))

            for i in range(len(matches_veiculos)):
                m = matches_veiculos[i]
                
                placa = m.group(1).replace('-', '')
                modelo = m.group(2).strip()
                km_atual = m.group(3)
                ano_fabr = m.group(4)

                start_idx = m.end()
                end_idx = matches_veiculos[i+1].start() if i+1 < len(matches_veiculos) else len(text)
                
                cabecalho_bruto = text[m.start():start_idx + 250].replace('\n', ' ')
                
                datas = re.findall(r'\d{2}/\d{2}/\d{4}', cabecalho_bruto)
                data_exec = datas[0] if len(datas) > 0 else ""
                data_inicio = datas[1] if len(datas) > 1 else ""
                data_fim = datas[2] if len(datas) > 2 else ""
                
                m_tempo = re.search(r'\b\d{2,3}:\d{2}\b', cabecalho_bruto)
                tempo_parado = m_tempo.group(0) if m_tempo else "00:00"
                
                resto_cab = cabecalho_bruto
                if data_fim: resto_cab = cabecalho_bruto[cabecalho_bruto.rfind(data_fim) + 10:]
                    
                valores = re.findall(r'[\d\.]+[,\.]\d{1,2}', resto_cab)
                hodometro = valores[0] if len(valores) > 0 else "0,00"
                total_mo = valores[1] if len(valores) > 1 else "0,00"
                total_pecas = valores[2] if len(valores) > 2 else "0,00"
                custo_im = valores[3] if len(valores) > 3 else "0,00"

                veiculo_info = {
                    'N. Veículo': placa,
                    'Modelo': modelo,
                    'Placa': placa,
                    'Km Atual': km_atual,
                    'Ano Fabr.': ano_fabr,
                    'Data Exec.': data_exec,
                    'Data Início': data_inicio,
                    'Data Fim': data_fim,
                    'Tempo Parado': tempo_parado,
                    'Hodômetro': hodometro,
                    'Total M-O': total_mo,
                    'Total Peças': total_pecas,
                    'Custo da IM': custo_im,
                    'Código da IM': '',
                    'Grupo veicular': 'DIFERENCIAL'
                }

                # Passo 2: O Segredo para as Quebras de Página (A Peneira Fina)
                bloco = text[start_idx:end_idx].strip()
                pecas_lines_raw = bloco.split('\n')
                pecas_lines = []
                
                # Lista de lixos que o PDF gera quando muda de página
                lixos_ignorar = [
                    "FLEETCOM - MANUTENÇÃO", "LINCE", "USUÁRIO:", "MANUTS. REALIZADAS", "PERÍODO:",
                    "N. VEÍCULO", "MODELO", "PLACA", "KM ATUAL", "ANO FABR", "DATA EXEC", "DATA INÍCIO", "DATA FIM", 
                    "TEMPO PARADO", "HODÔMETRO", "TOTAL M-O", "TOTAL PEÇAS", "CUSTO DA IM", "CÓDIGO DA IM", "GRUPO VEICULAR",
                    "QT CÓDIGO", "DESCRIÇÃO", "PR.TOT FORNECEDOR", "N. NF", "DESCONTOS", "GA.KMS", "GA. DIAS", "POSIÇÃO", "ORIGEM",
                    "PEÇAS", "VALOR", "ATUAL", "FABR.", "EXEC.", "INÍCIO", "INICIO", "FIM", "PARADO", "M-O", "DA IM", "KM ANO", "GAR.KMS", "GAR.DIAS", "DESCONTOS DESCRIÇÃO"
                ]
                
                for l in pecas_lines_raw:
                    l_upper = l.strip().upper()
                    is_lixo = any(l_upper.startswith(lixo) or l_upper == lixo for lixo in lixos_ignorar)
                    if not is_lixo and l.strip():
                        pecas_lines.append(l.strip())
                
                # Passo 3: Varredura dos Dados Limpos
                buffer_servico = []
                idx_line = 0
                
                while idx_line < len(pecas_lines):
                    line = pecas_lines[idx_line]
                    
                    # === MÃO DE OBRA ===
                    if "Fornecedor de Mão-de-Obra" in line:
                        desc_servico = " ".join(buffer_servico).strip()
                        
                        fornecedor = ""
                        nf = ""
                        valor = ""
                        desconto = ""
                        
                        # Olha as próximas linhas dinamicamente para caçar os valores fracionados pela página
                        offset = 1
                        while idx_line + offset < len(pecas_lines) and offset <= 6:
                            prox = pecas_lines[idx_line + offset]
                            
                            # Tenta ler tudo numa linha só (se não quebrou)
                            m_full = re.match(r'^(.*?)\s+(\d+)\s+([\d\.,]+)\s+([\d\.,]+)', prox)
                            if m_full:
                                fornecedor, nf, valor, desconto = m_full.groups()
                                offset += 1
                                break
                                
                            # Leitura vertical (Pedaço por pedaço após o corte do lixo da página)
                            if not fornecedor and not re.match(r'^[\d\.,]+$', prox):
                                fornecedor = prox
                            elif not nf and re.match(r'^\d+$', prox):
                                nf = prox
                            elif not valor and re.match(r'^[\d\.,]+$', prox):
                                valor = prox
                            elif not desconto and re.match(r'^[\d\.,]+$', prox):
                                desconto = prox
                                offset += 1
                                break 
                            
                            offset += 1
                            
                        # Atualiza o ponteiro para pular as linhas lidas
                        idx_line += (offset - 1)
                        
                        row = veiculo_info.copy()
                        row.update({
                            'Qt Código': '1', 'Código': '',
                            'Descrição': desc_servico if desc_servico else "Mão de Obra / Serviço",
                            'Pr.Tot Fornecedor': valor if valor else '0,00',
                            'Fornecedor': fornecedor.strip() if fornecedor else 'Não Informado',
                            'N. NF': nf,
                            'Descontos': desconto if desconto else '0,00',
                            'Ga.Kms': '0,00', 'Ga. Dias': '0,00', 'Posição': '', 'Origem': 'Mão de Obra'
                        })
                        parsed_data.append(row)
                        buffer_servico = []
                        
                        idx_line += 1
                        continue

                    # === PEÇAS ===
                    # Ajuste: Aceita NF com letras/traços ([A-Za-z0-9\-\/]+) e valores negativos ([\-\d\.,]+)
                    m_peca1 = re.match(r'^([\-\d\.,]+)\s+(.+?)\s+([\-\d\.,]+)\s+(.+?)\s+([A-Za-z0-9\-\/_]+)\s+([\-\d\.,]+)\s*(.*)$', line)
                    m_peca2 = re.match(r'^([\-\d\.,]+)\s+(.*?)\s+(.*?)\s+([\-\d\.,]+)\s+([A-Za-z0-9\-\/_]+)\s+([\-\d\.,]+)\s*(.*)$', line)
                    m_peca = m_peca1 if m_peca1 else m_peca2
                        
                    if m_peca:
                        # Descarrega o buffer se houver um serviço solto
                        if buffer_servico:
                            desc_tmp = " ".join(buffer_servico).strip()
                            if desc_tmp and len(desc_tmp) > 3:
                                row_mo = veiculo_info.copy()
                                row_mo.update({
                                    'Qt Código': '1', 'Código': '',
                                    'Descrição': desc_tmp,
                                    'Pr.Tot Fornecedor': '0,00',
                                    'Fornecedor': 'Serviço Interno / Sem NF',
                                    'N. NF': '', 'Descontos': '0,00', 'Ga.Kms': '0,00', 'Ga. Dias': '0,00', 'Posição': '', 'Origem': 'Interno'
                                })
                                parsed_data.append(row_mo)
                            buffer_servico = []

                        g3, g4 = m_peca.group(3).strip(), m_peca.group(4).strip()
                        
                        # Proteção contra números no fim da descrição
                        if re.match(r'^\-?\d+$', g3) and re.match(r'^[\-\d\.]+,\d{2}', g4):
                            valor = g4.split()[0]
                            fornecedor = g4.replace(valor, '').strip()
                            desc_full = m_peca.group(2).strip() + " " + g3
                        elif re.match(r'^[\-\d\.,]+$', g3):
                            valor, fornecedor = g3, g4
                            desc_full = m_peca.group(2).strip()
                        else:
                            valor, fornecedor = g4, g3
                            desc_full = m_peca.group(2).strip()

                        codigo = ""
                        m_cod = re.match(r'^([A-Z0-9\.\-]*\d[A-Z0-9\.\-]*)\s+(.*)', desc_full)
                        if m_cod:
                            codigo = m_cod.group(1)
                            desc_full = m_cod.group(2)

                        row = veiculo_info.copy()
                        row.update({
                            'Qt Código': m_peca.group(1),
                            'Código': codigo,
                            'Descrição': desc_full,
                            'Pr.Tot Fornecedor': valor,
                            'Fornecedor': fornecedor,
                            'N. NF': m_peca.group(5),
                            'Descontos': m_peca.group(6),
                            'Ga.Kms': '0,00', 'Ga. Dias': '0,00', 'Posição': '', 'Origem': m_peca.group(7).strip() if m_peca.group(7) else 'NF manutenção'
                        })
                        parsed_data.append(row)
                        idx_line += 1
                        continue

                    # === TEXTO SOLTO (DESCRIÇÃO DE SERVIÇO) ===
                    if re.match(r'^\d{2}:\d{2}\s+[\d\.,]+\s+[\d\.,]+', line):
                        idx_line += 1
                        continue

                    if not re.match(r'^\d{2}/\d{2}/\d{4}', line) and not re.match(r'^[\-\d\.,]+$', line) and len(line) > 3:
                        buffer_servico.append(line)
                        
                    idx_line += 1

                # Flush Final de serviços no fim do bloco
                if buffer_servico:
                    desc_tmp = " ".join(buffer_servico).strip()
                    if desc_tmp and len(desc_tmp) > 3:
                        row_mo = veiculo_info.copy()
                        row_mo.update({
                            'Qt Código': '1', 'Código': '',
                            'Descrição': desc_tmp,
                            'Pr.Tot Fornecedor': '0,00',
                            'Fornecedor': 'Serviço Interno / Sem NF',
                            'N. NF': '', 'Descontos': '0,00', 'Ga.Kms': '0,00', 'Ga. Dias': '0,00', 'Posição': '', 'Origem': 'Interno'
                        })
                        parsed_data.append(row_mo)

            if parsed_data:
                df_resultado = pd.DataFrame(parsed_data)
                
                colunas_moeda = ['Total M-O', 'Total Peças', 'Custo da IM', 'Pr.Tot Fornecedor', 'Descontos']
                for c in colunas_moeda:
                    if c in df_resultado.columns:
                        df_resultado[c] = pd.to_numeric(
                            df_resultado[c].astype(str).str.replace('.', '', regex=False).str.replace(',', '.', regex=False), 
                            errors='coerce'
                        ).fillna(0)
                        
                colunas_ordenadas = [
                    'N. Veículo', 'Modelo', 'Placa', 'Km Atual', 'Ano Fabr.', 
                    'Data Exec.', 'Data Início', 'Data Fim', 'Tempo Parado', 'Hodômetro', 
                    'Total M-O', 'Total Peças', 'Custo da IM', 'Código da IM', 'Grupo veicular',
                    'Qt Código', 'Código', 'Descrição', 'Pr.Tot Fornecedor', 'Fornecedor', 
                    'N. NF', 'Descontos', 'Ga.Kms', 'Ga. Dias', 'Posição', 'Origem'
                ]
                colunas_presentes = [c for c in colunas_ordenadas if c in df_resultado.columns]
                return df_resultado[colunas_presentes]
                
            return pd.DataFrame()

        if st.button("Processar Frota Completa", use_container_width=True):
            if not any([up_cam, up_car, up_mot]):
                st.warning("⚠️ Envie pelo menos um relatório em PDF para processar.")
            else:
                dict_dfs = {}
                if up_cam:
                    df = extrair_dados_fleetcom(up_cam)
                    if df is not None and not df.empty: dict_dfs["Caminhões"] = df
                if up_car:
                    df = extrair_dados_fleetcom(up_car)
                    if df is not None and not df.empty: dict_dfs["Carros"] = df
                if up_mot:
                    df = extrair_dados_fleetcom(up_mot)
                    if df is not None and not df.empty: dict_dfs["Motos"] = df

                if dict_dfs:
                    st.success("✅ Relatórios processados com sucesso!")
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                        formato_moeda = writer.book.add_format({'num_format': 'R$ #,##0.00'})
                        for nome_aba, df_sheet in dict_dfs.items():
                            df_sheet.to_excel(writer, index=False, sheet_name=nome_aba)
                            ws = writer.sheets[nome_aba]
                            for idx, col in enumerate(df_sheet.columns):
                                tamanho = max(df_sheet[col].astype(str).str.len().max(), len(col)) + 2
                                if col in ['Total M-O', 'Total Peças', 'Custo da IM', 'Pr.Tot Fornecedor', 'Descontos']:
                                    ws.set_column(idx, idx, 15, formato_moeda)
                                else:
                                    ws.set_column(idx, idx, min(tamanho, 45))
                    
                    output.seek(0)
                    st.download_button(
                        label="📥 Baixar Planilha Consolidada (Excel)",
                        data=output,
                        file_name="Manutencao_Frota_Completa.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary", use_container_width=True
                    )
                    
                    st.write("---")
                    st.write("### 👁️ Visualização dos Dados:")
                    for nome_aba, df_sheet in dict_dfs.items():
                        with st.expander(f"Frota: {nome_aba} ({len(df_sheet)} serviços)"):
                            st.dataframe(df_sheet, use_container_width=True)
                else:
                    st.error("Não foi possível extrair dados dos arquivos. Verifique o padrão do PDF.")
# ====================================================================
# 6. SETOR COMERCIAL
# ====================================================================
def commercial_page():
    st.title("Setor Comercial")
    
    col_voltar, col_vazio = st.columns([1, 5])
    with col_voltar:
        if st.button("⬅️ Voltar"):
            st.session_state['current_page'] = 'home'
            st.rerun()

    st.markdown("---")
    script_selection = st.selectbox(
        "Selecione o script:", 
        ("Selecione...", "Troca de Canal", "Circuito Execução", "Planejamento Estratégico", "Limite de Credito", "Plano de Market Share"), 
        key="com_select"
    )

    # --- SCRIPT 1: TROCA DE CANAL ---
    if script_selection == "Troca de Canal":
        st.subheader("Troca de Canal")
        
        def transform_google_forms_data(df):
            processed_records = []
            if df.empty or len(df.columns) < 28: return pd.DataFrame()
            for index, row in df.iterrows():
                if not isinstance(row, pd.Series) or len(row) < 28: continue
                try:
                    data_value = row.iloc[0]; sv_value = row.iloc[1]
                    vd_consolidated_parts = [str(row.iloc[col_idx]).strip() for col_idx in range(2, min(5, len(row))) if pd.notna(row.iloc[col_idx])]
                    vd_final = ' | '.join(vd_consolidated_parts) if vd_consolidated_parts else None
                    para_value = row.iloc[27]
                    for col_idx in range(5, min(27, len(row))):
                        cell_content = str(row.iloc[col_idx]).strip()
                        if not cell_content or cell_content.lower() == 'nan': continue
                        de_category_match = re.search(r'\((.*?)\)', cell_content)
                        de_category_val = de_category_match.group(1).strip() if de_category_match else None
                        pdv_info_raw = re.sub(r'\s*\([^)]*\)\s*$', '', cell_content).strip()
                        pdv_info_val = re.sub(r'^\s*(?:\b\w+\s+)?\d+\s*[\|-]\s*', '', pdv_info_raw, 1).strip() if pdv_info_raw else None
                        if pdv_info_val or de_category_val:
                            processed_records.append({'DATA': data_value, 'SV': sv_value, 'VD': vd_final, 'PDV': pdv_info_val, 'DE': de_category_val, 'PARA': para_value, 'Status': ''})
                except IndexError: continue
            return pd.DataFrame(processed_records)

        uploaded_file_1 = st.file_uploader("Envie o arquivo (.xlsx)", type=["xlsx"], key="troca_canal_uploader") 
        if uploaded_file_1 is not None:
            try:
                df_forms = pd.read_excel(uploaded_file_1)
                st.dataframe(df_forms.head())
                final_df_forms = transform_google_forms_data(df_forms)
                if not final_df_forms.empty:
                    output = io.BytesIO(); final_df_forms.to_excel(output, index=False); output.seek(0)
                    workbook = load_workbook(output); sheet = workbook.active
                    dv = DataValidation(type="list", formula1='"Aprovado,Não Aprovado"', allow_blank=True)
                    try:
                        col_letter = get_column_letter(final_df_forms.columns.get_loc('Status') + 1)
                        dv.add(f'{col_letter}2:{col_letter}{sheet.max_row}'); sheet.add_data_validation(dv)
                    except KeyError: pass
                    output_with_dropdown = io.BytesIO(); workbook.save(output_with_dropdown); output_with_dropdown.seek(0)
                    st.dataframe(final_df_forms)
                    st.download_button(label="📥 Baixar Arquivo", data=output_with_dropdown.getvalue(), file_name="troca_canal_processada.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            except Exception as e: st.error(f"Erro: {e}")

    # --- SCRIPT 2: CIRCUITO EXECUÇÃO E COM12 ---
    elif script_selection == "Circuito Execução":
        st.subheader("Circuito Execução")

        def transform_points_columns(df):
            df_transformed = df.copy()
            
            colunas_remover = [col for col in df_transformed.columns if str(col).strip().upper() in ['PONTUAÇÃO', 'PONTUACAO']]
            if colunas_remover:
                df_transformed.drop(columns=colunas_remover, inplace=True, errors='ignore')

            header_pattern = re.compile(r"\(\s*(\d+)\s*Pontos\s*\)", re.IGNORECASE)
            cell_pattern = re.compile(r"\(\s*(\d+)\s*Pontos\s*\)", re.IGNORECASE)

            for col in df_transformed.columns:
                str_col = str(col)
                header_match = header_pattern.search(str_col)
                
                default_points = int(header_match.group(1)) if header_match else None
                
                if header_match or "PRECIFICADAS" in str_col.upper():
                    
                    def process_cell(val):
                        s = str(val).strip()
                        s_upper = s.upper()
                        
                        cell_match = cell_pattern.search(s)
                        if cell_match:
                            return int(cell_match.group(1))
                        
                        if default_points is not None:
                            palavras_chave = ["SIM", "PRESENÇA", "PRESENCA", "OK", "CONFORME", "VISIBILIDADE"]
                            if any(x in s_upper for x in palavras_chave):
                                return default_points
                            if s == '1': 
                                return default_points
                        
                        return 0

                    df_transformed[col] = df_transformed[col].apply(process_cell)
            
            cols_presenca = [c for c in df_transformed.columns if str(c).strip().upper().startswith("PRESENÇA")]
            df_transformed["PRESENÇA"] = df_transformed[cols_presenca].apply(pd.to_numeric, errors='coerce').sum(axis=1)

            cols_visibilidade = [c for c in df_transformed.columns if str(c).strip().upper().startswith("VISIBILIDADE")]
            df_transformed["VISIBILIDADE"] = df_transformed[cols_visibilidade].apply(pd.to_numeric, errors='coerce').sum(axis=1)

            cols_posicionamento = [c for c in df_transformed.columns if str(c).strip().upper().startswith("POSICIONAMENTO DE NOSSO PRODUTOS")]
            df_transformed["POSICIONAMENTO DE NOSSO PRODUTOS"] = df_transformed[cols_posicionamento].apply(pd.to_numeric, errors='coerce').sum(axis=1)

            cols_geladas = [c for c in df_transformed.columns if str(c).strip().upper().startswith("TEM NOSSAS CERVEJAS GELADAS")]
            df_transformed["TEM NOSSAS CERVEJAS GELADAS"] = df_transformed[cols_geladas].apply(pd.to_numeric, errors='coerce').sum(axis=1)
            
            cols_precificadas = [c for c in df_transformed.columns if "PRECIFICADAS" in str(c).upper()]
            if cols_precificadas:
                df_transformed["TODAS AS NOSSAS CERVEJAS ESTÃO PRECIFICADAS"] = df_transformed[cols_precificadas].apply(pd.to_numeric, errors='coerce').sum(axis=1)
            else:
                df_transformed["TODAS AS NOSSAS CERVEJAS ESTÃO PRECIFICADAS"] = 0

            pontuacao_total = (
                df_transformed["PRESENÇA"] + 
                df_transformed["VISIBILIDADE"] + 
                df_transformed["POSICIONAMENTO DE NOSSO PRODUTOS"] + 
                df_transformed["TEM NOSSAS CERVEJAS GELADAS"] +
                df_transformed["TODAS AS NOSSAS CERVEJAS ESTÃO PRECIFICADAS"]
            )
            
            porcentagem = pontuacao_total / 400.0
            
            df_transformed.insert(1, '% de Pontuação', porcentagem)
            
            return df_transformed

        uploaded_file_2 = st.file_uploader("Envie o arquivo do Circuito (.xlsx)", type=["xlsx"], key="circuito_exec_uploader") 
        if uploaded_file_2 is not None:
            try:
                df_points = pd.read_excel(uploaded_file_2)
                st.write("Visualização original Circuito (5 primeiras linhas):")
                st.dataframe(df_points.head())
                
                df_transformed = transform_points_columns(df_points)
                
                st.success("Transformação do Circuito concluída!")
                st.write("Visualização processada Circuito:")
                st.dataframe(df_transformed.head())
                
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
                    df_transformed.to_excel(writer, index=False)
                output.seek(0)
                
                st.download_button(
                    label="📥 Baixar Arquivo Circuito Transformado",
                    data=output,
                    file_name="circuito_execucao_final.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                st.error(f"Erro ao processar o arquivo de Circuito: {e}")

        st.markdown("---")
        st.subheader("Transformação e Agrupamento COM12")
        st.info("Deixa apenas 1 linha por **CodCli**, somando as métricas, adicionando colunas (primeiro SKUs, depois HLs) e removendo colunas indesejadas.")

        def transform_com12_data(df):
            df_transformed = df.copy()
            
            cols_to_sum = ['QtdeVdaSemBonifTOTAL', 'BonRevenda', 'BonFabrica', 'QtdeVdaSemBonRGB', 'BonRevRGB', 'BonFabRGB', 'consideraSkuTotal', 'HL', 'HL RGB']
            
            for col in cols_to_sum:
                if col in df_transformed.columns:
                    df_transformed[col] = df_transformed[col].astype(str).str.replace(',', '.', regex=False).str.strip()
                    df_transformed[col] = df_transformed[col].replace(['-', ''], '0')
                    df_transformed[col] = pd.to_numeric(df_transformed[col], errors='coerce').fillna(0)
                    
            df_transformed['TotalVda'] = df_transformed.get('QtdeVdaSemBonifTOTAL', 0) + df_transformed.get('BonRevenda', 0) + df_transformed.get('BonFabrica', 0)
            df_transformed['TotalVdaRGB'] = df_transformed.get('QtdeVdaSemBonRGB', 0) + df_transformed.get('BonRevRGB', 0) + df_transformed.get('BonFabRGB', 0)
            
            cols_to_sum.extend(['TotalVda', 'TotalVdaRGB'])

            if 'RefMes' in df_transformed.columns:
                df_transformed['RefMes'] = pd.to_datetime(df_transformed['RefMes'], errors='coerce').dt.strftime('%m/%Y')

            if 'Vend Cli (Cód)' in df_transformed.columns:
                df_transformed['Vend Cli (Cód)'] = df_transformed['Vend Cli (Cód)'].astype(str).str.replace('2216-', '', regex=False)
            if 'Sup Cli (Cód)' in df_transformed.columns:
                df_transformed['Sup Cli (Cód)'] = df_transformed['Sup Cli (Cód)'].astype(str).str.replace('2216-', '', regex=False)
                    
            pivot_meses = pd.DataFrame()
            
            if len(df.columns) > 11:
                col_mes = df.columns[7]
                col_sku = df.columns[10]
                col_valor = df.columns[11]
                
                if col_mes in df_transformed.columns:
                    if col_valor in df_transformed.columns:
                        df_transformed['TEMP_VALOR_L'] = df_transformed[col_valor].astype(str).str.replace(',', '.', regex=False).str.replace('-', '0')
                        df_transformed['TEMP_VALOR_L'] = pd.to_numeric(df_transformed['TEMP_VALOR_L'], errors='coerce').fillna(0)
                    else:
                        df_transformed['TEMP_VALOR_L'] = 0
                        
                    pivot_l = pd.pivot_table(
                        df_transformed,
                        values='TEMP_VALOR_L',
                        index='CodCli',
                        columns=col_mes,
                        aggfunc='sum',
                        fill_value=0
                    )
                    pivot_l.columns = [f"{str(c).strip()} (HL)" for c in pivot_l.columns]
                    pivot_l = pivot_l.reset_index()

                    if col_sku in df_transformed.columns:
                        df_transformed['TEMP_VALOR_K'] = df_transformed[col_sku].astype(str).str.replace(',', '.', regex=False).str.replace('-', '0')
                        df_transformed['TEMP_VALOR_K'] = pd.to_numeric(df_transformed['TEMP_VALOR_K'], errors='coerce').fillna(0)
                    else:
                        df_transformed['TEMP_VALOR_K'] = 0

                    pivot_k = pd.pivot_table(
                        df_transformed,
                        values='TEMP_VALOR_K',
                        index='CodCli',
                        columns=col_mes,
                        aggfunc='sum',
                        fill_value=0
                    )
                    pivot_k.columns = [f"{str(c).strip()} (SKU)" for c in pivot_k.columns]
                    pivot_k = pivot_k.reset_index()

                    pivot_meses = pd.merge(pivot_l, pivot_k, on='CodCli', how='outer')
                    
                    meses_unicos = [str(m).strip() for m in df_transformed[col_mes].dropna().unique()]
                    
                    colunas_ordenadas = ['CodCli']
                    
                    for m in meses_unicos:
                        col_sku_str = f"{m} (SKU)"
                        if col_sku_str in pivot_meses.columns:
                            colunas_ordenadas.append(col_sku_str)
                            
                    for m in meses_unicos:
                        col_hl = f"{m} (HL)"
                        if col_hl in pivot_meses.columns:
                            colunas_ordenadas.append(col_hl)
                    
                    cols_restantes = [c for c in pivot_meses.columns if c not in colunas_ordenadas]
                    pivot_meses = pivot_meses[colunas_ordenadas + cols_restantes]

                    df_transformed.drop(columns=['TEMP_VALOR_L', 'TEMP_VALOR_K'], inplace=True, errors='ignore')

            agg_dict = {}
            for col in df_transformed.columns:
                if col == 'CodCli':
                    continue
                elif col in cols_to_sum:
                    agg_dict[col] = 'sum'
                elif col in ['ProdCod', 'ProdDesc']:
                    agg_dict[col] = lambda x: ', '.join(x.dropna().astype(str).unique())
                else:
                    agg_dict[col] = 'first'
                    
            df_grouped = df_transformed.groupby('CodCli', as_index=False).agg(agg_dict)

            if not pivot_meses.empty:
                df_grouped = pd.merge(df_grouped, pivot_meses, on='CodCli', how='left')
                meses_adicionados = [c for c in pivot_meses.columns if c != 'CodCli']
                df_grouped[meses_adicionados] = df_grouped[meses_adicionados].fillna(0)

            colunas_para_remover = ['HL RGB', 'TotalVda', 'TotalVdaRGB', 'RefMes']
            df_grouped.drop(columns=[c for c in colunas_para_remover if c in df_grouped.columns], inplace=True)

            return df_grouped

        uploaded_com12 = st.file_uploader("Envie o arquivo COM12 (.xlsx ou .csv)", type=["xlsx", "csv"], key="com12_uploader")
        if uploaded_com12 is not None:
            try:
                if uploaded_com12.name.endswith('.csv'):
                    df_com12 = pd.read_csv(uploaded_com12)
                else:
                    df_com12 = pd.read_excel(uploaded_com12)
                    
                st.write("Visualização original COM12 (5 primeiras linhas):")
                st.dataframe(df_com12.head())
                
                df_com12_grouped = transform_com12_data(df_com12)
                
                st.success("Agrupamento COM12 concluído!")
                st.write("Visualização processada COM12:")
                st.dataframe(df_com12_grouped.head())
                
                output_com12 = io.BytesIO()
                with pd.ExcelWriter(output_com12, engine="xlsxwriter") as writer:
                    df_com12_grouped.to_excel(writer, index=False)
                output_com12.seek(0)
                
                st.download_button(
                    label="📥 Baixar Arquivo COM12 Agrupado",
                    data=output_com12,
                    file_name="COM12_Agrupado.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                st.error(f"Erro ao processar o arquivo COM12: {e}")

    # --- SCRIPT 3: PLANEJAMENTO ESTRATÉGICO ---
    elif script_selection == "Planejamento Estratégico":
        st.subheader("Planejamento Estratégico (Diamante e Ouro & 50 VOLHNK)")
        st.info("O arquivo original deve conter as colunas: CodCli, Razão Social, RefMes, xPorte, QtdeSaidaHL, QtdSaidaHLRGB e ConsideraSKUTOTAL.")
        
        uploaded_pe = st.file_uploader("Envie o arquivo (.xlsx ou .csv)", type=["xlsx", "csv"], key="pe_uploader")
        
        if uploaded_pe is not None:
            try:
                if uploaded_pe.name.endswith('.csv'):
                    df_pe = pd.read_csv(uploaded_pe)
                else:
                    df_pe = pd.read_excel(uploaded_pe)
                
                st.write("Visualização original (5 primeiras linhas):")
                st.dataframe(df_pe.head())
                
                df_pe['RefMes'] = pd.to_datetime(df_pe['RefMes'], errors='coerce')
                mes_atual = df_pe['RefMes'].max()
                
                if pd.isna(mes_atual):
                    st.error("Erro ao identificar as datas na coluna 'RefMes'.")
                else:
                    ano_anterior = mes_atual.year - 1
                    mes_atual_ano_anterior = mes_atual - pd.DateOffset(years=1)
                    
                    df_diamante = df_pe.copy()
                    
                    if 'xPorte' in df_diamante.columns:
                        df_diamante['xPorte'] = df_diamante['xPorte'].astype(str).str.strip().str.upper()
                        df_diamante = df_diamante[df_diamante['xPorte'].isin(['O', 'D'])].copy()
                    else:
                        st.warning("Coluna 'xPorte' não encontrada na aba Diamante e Ouro. O filtro não será aplicado.")
                    
                    meses_3m = [mes_atual - pd.DateOffset(months=i) for i in [1, 2, 3]]
                    
                    df_atual_diamante = df_diamante[df_diamante['RefMes'] == mes_atual].groupby('CodCli').agg({
                        'QtdeSaidaHL': 'sum',
                        'QtdSaidaHLRGB': 'sum',
                        'ConsideraSKUTOTAL': 'sum'
                    }).reset_index().rename(columns={
                        'QtdeSaidaHL': 'HL_ATUAL',
                        'QtdSaidaHLRGB': 'HLRGB_ATUAL',
                        'ConsideraSKUTOTAL': 'SKUS_ATUAL'
                    })
                    
                    df_3m = df_diamante[df_diamante['RefMes'].isin(meses_3m)].groupby('CodCli').agg({
                        'QtdeSaidaHL': 'sum',
                        'QtdSaidaHLRGB': 'sum',
                        'ConsideraSKUTOTAL': 'sum'
                    }).reset_index()
                    
                    df_3m['HL_3M'] = df_3m['QtdeSaidaHL'] / 3
                    df_3m['HLRGB_3M'] = df_3m['QtdSaidaHLRGB'] / 3
                    df_3m['SKUS_3M'] = df_3m['ConsideraSKUTOTAL'] / 3
                    df_3m.drop(columns=['QtdeSaidaHL', 'QtdSaidaHLRGB', 'ConsideraSKUTOTAL'], inplace=True)
                    
                    cols_base = ['CodCli', 'Razão Social', 'SV Cód', 'VD Cód', 'xPorte']
                    cols_base = [c for c in cols_base if c in df_diamante.columns]
                    
                    df_base_diamante = df_diamante.sort_values('RefMes').drop_duplicates('CodCli', keep='last')[cols_base]
                    
                    df_final_diamante = df_base_diamante.merge(df_3m, on='CodCli', how='left').merge(df_atual_diamante, on='CodCli', how='left').fillna(0)
                    
                    def get_status(atual, media):
                        if atual == 0 and media == 0:
                            return 'INTRODUZIR'
                        elif atual >= media:
                            return 'PROTEGER'
                        else:
                            return 'ATACAR'
                    
                    df_final_diamante['STATUS_HL'] = df_final_diamante.apply(lambda r: get_status(r['HL_ATUAL'], r['HL_3M']), axis=1)
                    df_final_diamante['STATUS_HLRGB'] = df_final_diamante.apply(lambda r: get_status(r['HLRGB_ATUAL'], r['HLRGB_3M']), axis=1)
                    df_final_diamante['STATUS_SKUS'] = df_final_diamante.apply(lambda r: get_status(r['SKUS_ATUAL'], r['SKUS_3M']), axis=1)
                    
                    def get_acao(atual, media):
                        if atual == 0:
                            return 'PDV SEM COBERTURA'
                        elif atual < media:
                            return 'PDV EM QUEDA'
                        else:
                            return ''
                    
                    df_final_diamante['AÇÃO'] = df_final_diamante.apply(lambda r: get_acao(r['HL_ATUAL'], r['HL_3M']), axis=1)
                    df_final_diamante['Plano de Ação'] = ''
                    
                    col_order_diamante = cols_base + [
                        'HL_3M', 'HL_ATUAL', 'STATUS_HL', 
                        'HLRGB_3M', 'HLRGB_ATUAL', 'STATUS_HLRGB', 
                        'SKUS_3M', 'SKUS_ATUAL', 'STATUS_SKUS', 
                        'AÇÃO', 'Plano de Ação'
                    ]
                    df_final_diamante = df_final_diamante[col_order_diamante]

                    df_50 = df_pe.copy()
                    
                    df_ly = df_50[df_50['RefMes'].dt.year == ano_anterior].groupby('CodCli').agg({
                        'QtdeSaidaHL': 'sum',
                        'QtdSaidaHLRGB': 'sum'
                    }).reset_index().rename(columns={
                        'QtdeSaidaHL': 'SellOut_Total_LY',
                        'QtdSaidaHLRGB': 'SellOut_RGB_LY'
                    })
                    
                    df_meta = df_50[df_50['RefMes'] == mes_atual_ano_anterior].groupby('CodCli').agg({
                        'QtdeSaidaHL': 'sum',
                        'QtdSaidaHLRGB': 'sum'
                    }).reset_index()
                    df_meta['Meta_SellOut_Total'] = df_meta['QtdeSaidaHL'] * 1.05
                    df_meta['Meta_SellOut_RGB'] = df_meta['QtdSaidaHLRGB'] * 1.05
                    df_meta.drop(columns=['QtdeSaidaHL', 'QtdSaidaHLRGB'], inplace=True)
                    
                    df_atual_50 = df_50[df_50['RefMes'] == mes_atual].groupby('CodCli').agg({
                        'QtdeSaidaHL': 'sum',
                        'QtdSaidaHLRGB': 'sum'
                    }).reset_index().rename(columns={
                        'QtdeSaidaHL': 'SellOut_Total_Atual',
                        'QtdSaidaHLRGB': 'SellOut_RGB_Atual'
                    })
                    
                    cols_base_50 = [c for c in cols_base if c in df_50.columns]
                    df_base_50 = df_50.sort_values('RefMes').drop_duplicates('CodCli', keep='last')[cols_base_50]
                    
                    df_final_50 = df_base_50.merge(df_ly, on='CodCli', how='left') \
                                            .merge(df_meta, on='CodCli', how='left') \
                                            .merge(df_atual_50, on='CodCli', how='left').fillna(0)
                    
                    df_final_50.loc[df_final_50['Meta_SellOut_Total'] == 0, 'Meta_SellOut_Total'] = 1
                    df_final_50.loc[df_final_50['Meta_SellOut_RGB'] == 0, 'Meta_SellOut_RGB'] = 1
                    
                    df_final_50['SellOut_Total_%'] = np.where(
                        df_final_50['Meta_SellOut_Total'] > 0, 
                        df_final_50['SellOut_Total_Atual'] / df_final_50['Meta_SellOut_Total'], 
                        0
                    )
                    df_final_50['SellOut_RGB_%'] = np.where(
                        df_final_50['Meta_SellOut_RGB'] > 0, 
                        df_final_50['SellOut_RGB_Atual'] / df_final_50['Meta_SellOut_RGB'], 
                        0
                    )
                    
                    col_order_50 = cols_base_50 + [
                        'SellOut_Total_LY', 'Meta_SellOut_Total', 'SellOut_Total_Atual', 'SellOut_Total_%',
                        'SellOut_RGB_LY', 'Meta_SellOut_RGB', 'SellOut_RGB_Atual', 'SellOut_RGB_%'
                    ]
                    df_final_50 = df_final_50[col_order_50]

                    st.success("Cálculos do Planejamento Estratégico concluídos com sucesso!")
                    
                    st.write("**Resumo da Aba: DIAMANTE E OURO (xPorte O e D)**")
                    st.dataframe(df_final_diamante.head())
                    
                    st.write("**Resumo da Aba: 50 VOLHNK (Geral)**")
                    st.dataframe(df_final_50.head())
                    
                    output_pe = io.BytesIO()
                    with pd.ExcelWriter(output_pe, engine="xlsxwriter") as writer:
                        df_final_diamante.to_excel(writer, sheet_name="DIAMANTE E OURO", index=False)
                        df_final_50.to_excel(writer, sheet_name="50 VOLHNK", index=False)
                    output_pe.seek(0)
                    
                    st.download_button(
                        label="📥 Baixar Planejamento Estratégico Completo",
                        data=output_pe,
                        file_name="Planejamento_Estrategico_Completo.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
            except Exception as e:
                st.error(f"Erro ao processar o arquivo de Planejamento Estratégico: {e}")

    # --- SCRIPT 4: LIMITE DE CRÉDITO ---
    elif script_selection == "Limite de Credito":
        st.subheader("Análise para Limite de Crédito (Faturamento por Mês)")
        st.info("O arquivo deve conter as colunas: CodCli, Fantasia, RefMes, Faturamento e xPorte.")
        
        uploaded_lc = st.file_uploader("Envie o arquivo de Vendas (.xlsx ou .csv)", type=["xlsx", "csv"], key="lc_uploader")
        
        if uploaded_lc is not None:
            try:
                st.info("Lendo e processando os dados...")
                if uploaded_lc.name.endswith('.csv'):
                    df_lc = pd.read_csv(uploaded_lc)
                else:
                    df_lc = pd.read_excel(uploaded_lc)
                
                df_lc['RefMes'] = pd.to_datetime(df_lc['RefMes'], errors='coerce')
                df_lc['MesAno'] = df_lc['RefMes'].dt.strftime('%m/%Y') 
                
                if 'Faturamento' in df_lc.columns:
                    df_lc['Faturamento'] = pd.to_numeric(df_lc['Faturamento'], errors='coerce').fillna(0)
                else:
                    st.error("A coluna 'Faturamento' não foi encontrada no arquivo!")
                    st.stop()
                
                if 'xPorte' in df_lc.columns:
                    df_lc['xPorte'] = df_lc['xPorte'].astype(str).str.strip().str.upper()
                    map_porte = {'O': 'OURO', 'D': 'DIAMANTE', 'P': 'PRATA', 'B': 'BRONZE'}
                    df_lc['xPorte'] = df_lc['xPorte'].map(map_porte).fillna(df_lc['xPorte'])
                
                colunas_indice = ['CodCli', 'Fantasia', 'VD', 'SV', 'GerPedido', 'xPorte']
                colunas_indice_existentes = [col for col in colunas_indice if col in df_lc.columns]
                
                if not colunas_indice_existentes or 'CodCli' not in colunas_indice_existentes:
                     st.error("A coluna 'CodCli' não foi encontrada para identificação do cliente.")
                     st.stop()

                df_cadastral = df_lc[colunas_indice_existentes].drop_duplicates(subset=['CodCli'], keep='last')

                df_faturamento = df_lc.groupby(['CodCli', 'MesAno'])['Faturamento'].sum().reset_index()
                df_pivot_valores = df_faturamento.pivot(index='CodCli', columns='MesAno', values='Faturamento').fillna(0).reset_index()
                
                df_pivot_lc = pd.merge(df_cadastral, df_pivot_valores, on='CodCli', how='left')
                
                meses_cols = [col for col in df_pivot_valores.columns if col != 'CodCli']
                
                meses_ordenados = sorted(meses_cols, key=lambda x: datetime.strptime(x, '%m/%Y'))
                ultimos_3_meses = meses_ordenados[-3:] if len(meses_ordenados) >= 3 else meses_ordenados
                
                num_meses = len(ultimos_3_meses)
                df_pivot_lc['Media 3 Meses'] = (df_pivot_lc[ultimos_3_meses].sum(axis=1) / num_meses) if num_meses > 0 else 0
                
                def calcular_limite(row):
                    porte = str(row.get('xPorte', '')).strip().upper()
                    media = row['Media 3 Meses']
                    
                    if porte in ['DIAMANTE', 'D']:
                        return max(media * 1.50, 1000.0)
                    elif porte in ['OURO', 'O']:
                        return max(media * 1.40, 600.0)
                    elif porte in ['PRATA', 'P']:
                        return max(media * 1.30, 400.0)
                    elif porte in ['BRONZE', 'B']:
                        return max(media * 1.20, 200.0)
                    else:
                        return 0.0
                        
                df_pivot_lc['Limite de Credito'] = df_pivot_lc.apply(calcular_limite, axis=1).round(2)
                df_pivot_lc['Media 3 Meses'] = df_pivot_lc['Media 3 Meses'].round(2)
                
                df_pivot_lc['Faturamento Total'] = df_pivot_lc[meses_cols].sum(axis=1).round(2)
                cols_finais = colunas_indice_existentes + meses_ordenados + ['Faturamento Total', 'Media 3 Meses', 'Limite de Credito']
                df_pivot_lc = df_pivot_lc[cols_finais]
                
                st.success("Tabela processada com sucesso!")
                st.write("**Resumo - Limite de Crédito:**")
                st.dataframe(df_pivot_lc.head(15))
                
                output_lc = io.BytesIO()
                with pd.ExcelWriter(output_lc, engine="xlsxwriter") as writer:
                    df_pivot_lc.to_excel(writer, sheet_name="Limite de Credito", index=False)
                output_lc.seek(0)
                
                st.download_button(
                    label="📥 Baixar Análise de Limite de Crédito",
                    data=output_lc,
                    file_name="Limite_Credito_Analisado.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                st.error(f"Erro ao processar o arquivo de Limite de Crédito: {e}")

    # =============================================================
    # NOVO SCRIPT 5: PLANO DE MARKET SHARE (VOLUME & EXECUÇÃO SEPARADOS)
    # =============================================================
    elif script_selection == "Plano de Market Share":
        st.subheader("Plano de Market Share (Arquivos Independentes)")
        st.info("Pode enviar a base de Volume, a base de Execução, ou ambas. O sistema irá gerar ficheiros Excel separados para cada uma.")
        
        col1, col2 = st.columns(2)
        with col1:
            uploaded_ms = st.file_uploader("1. Envie a base de Volume (RGB, MAINSTREAM...)", type=["xlsx", "csv"], key="ms_uploader")
        with col2:
            uploaded_exec = st.file_uploader("2. Envie a base de Execução (SKU, mixRGB, Drop)", type=["xlsx", "csv"], key="exec_uploader")
        
        meses_pt = {1: 'jan', 2: 'fev', 3: 'mar', 4: 'abr', 5: 'mai', 6: 'jun', 
                    7: 'jul', 8: 'ago', 9: 'set', 10: 'out', 11: 'nov', 12: 'dez'}
        
        # LISTA ATUALIZADA: Define a ORDEM EXATA em que as colunas vão aparecer no Excel final
        colunas_cadastrais = [
            'VendCliCod', 'Vend Cli (Cód)', 
            'SupCliCod', 'Sup Cli (Cód)', 
            'RazaoSocial', 'Razão Social', 
            'Fantasia', 'CNPJ Cli', 'Cidade', 
            'CanalCod', 'Canal', 'Porte', 
            'PastaCliCod', 'Pasta Cli'
        ]

        # ==============================================================
        # PROCESSAMENTO 1: VOLUME (RGB, MAINSTREAM, PREMIUM)
        # ==============================================================
        if uploaded_ms is not None:
            try:
                st.markdown("### 📊 Processamento de Volume")
                df_vol = pd.read_csv(uploaded_ms) if uploaded_ms.name.endswith('.csv') else pd.read_excel(uploaded_ms)
                
                if 'RefMes' not in df_vol.columns or 'CodCli' not in df_vol.columns:
                    st.error("As colunas 'RefMes' e 'CodCli' são obrigatórias na base de Volume.")
                else:
                    df_vol['RefMes'] = pd.to_datetime(df_vol['RefMes'], errors='coerce')
                    df_vol = df_vol.dropna(subset=['RefMes'])
                    df_vol['MesFormatado'] = df_vol['RefMes'].dt.month.map(meses_pt) + '/' + df_vol['RefMes'].dt.strftime('%y')
                    df_vol['Trimestre'] = df_vol['RefMes'].dt.quarter
                    df_vol['Ano'] = df_vol['RefMes'].dt.year
                    
                    metricas_vol = ['RGB', 'MAINSTREAM', 'PREMIUM']
                    for col in metricas_vol:
                        if col in df_vol.columns:
                            df_vol[col] = df_vol[col].astype(str).str.replace(',', '.', regex=False)
                            df_vol[col] = pd.to_numeric(df_vol[col], errors='coerce').fillna(0)
                        else:
                            df_vol[col] = 0.0
                            
                    col_cad_exist_vol = [c for c in colunas_cadastrais if c in df_vol.columns]
                    df_cad_vol = df_vol[['CodCli'] + col_cad_exist_vol].drop_duplicates(subset=['CodCli'], keep='last')
                    
                    # Metas 2025 -> 2026
                    df_2025_vol = df_vol[df_vol['Ano'] == 2025].copy()
                    df_metas_vol = pd.DataFrame({'CodCli': df_cad_vol['CodCli'].unique()})
                    
                    if not df_2025_vol.empty:
                        df_tri_25 = df_2025_vol.groupby(['CodCli', 'Trimestre'])[metricas_vol].sum().unstack(fill_value=0)
                        df_tri_25.columns = [f"Q{q}_{met}" for met, q in df_tri_25.columns]
                        df_tri_25 = df_tri_25.reset_index()
                        for met in metricas_vol:
                            for q in [1, 2, 3, 4]:
                                if f"Q{q}_{met}" not in df_tri_25.columns: df_tri_25[f"Q{q}_{met}"] = 0.0
                                    
                        def calc_meta_vol(row, met):
                            q1, q2, q3, q4 = row[f'Q1_{met}'], row[f'Q2_{met}'], row[f'Q3_{met}'], row[f'Q4_{met}']
                            ultimo = q4 if q4 > 0 else (q3 if q3 > 0 else (q2 if q2 > 0 else (q1 if q1 > 0 else 0)))
                            b1 = q1 if q1 > 0 else ultimo
                            b2 = q2 if q2 > 0 else b1
                            b3 = q3 if q3 > 0 else b2
                            b4 = q4 if q4 > 0 else b3
                            return pd.Series([b1*1.10, b2*1.10, b3*1.10, b4*1.10])
                            
                        for met in metricas_vol:
                            cols = [f'Meta {q}° Tri ({met})' for q in [1, 2, 3, 4]]
                            df_tri_25[cols] = df_tri_25.apply(lambda r: calc_meta_vol(r, met), axis=1)
                            
                        cols_merge = ['CodCli'] + [f'Meta {q}° Tri ({met})' for met in metricas_vol for q in [1, 2, 3, 4]]
                        df_metas_vol = pd.merge(df_metas_vol, df_tri_25[cols_merge], on='CodCli', how='left').fillna(0)
                    else:
                        for met in metricas_vol:
                            for q in [1, 2, 3, 4]: df_metas_vol[f'Meta {q}° Tri ({met})'] = 0.0

                    # Pivot
                    df_pivot_vol = df_vol.pivot_table(index='CodCli', columns='MesFormatado', values=metricas_vol, aggfunc='sum', fill_value=0)
                    df_pivot_vol.columns = [f"{mes} ({'RGB' if met.upper()=='RGB' else met.upper()})" for met, mes in df_pivot_vol.columns]
                    df_pivot_vol = df_pivot_vol.reset_index()
                    
                    df_final_vol = pd.merge(df_cad_vol, df_pivot_vol, on='CodCli', how='left').fillna(0)
                    df_final_vol = pd.merge(df_final_vol, df_metas_vol, on='CodCli', how='left').fillna(0)

                    # Ordenação e Realizado
                    anos_unicos = sorted(df_vol['Ano'].unique())
                    meses_26_pres = df_vol[df_vol['Ano'] == 2026]['RefMes'].dt.month.unique() if 2026 in anos_unicos else []
                    cols_finais = ['CodCli'] + col_cad_exist_vol
                    
                    for met in metricas_vol:
                        nome_met = 'RGB' if met.upper() == 'RGB' else met.upper()
                        for ano in anos_unicos:
                            ano_str = str(ano)[-2:]
                            if ano < 2026:
                                for m_num in range(1, 13):
                                    col_m = f"{meses_pt[m_num]}/{ano_str} ({nome_met})"
                                    if col_m in df_final_vol.columns: cols_finais.append(col_m)
                            elif ano == 2026:
                                for q in [1, 2, 3, 4]:
                                    meses_tri = [(q-1)*3 + 1, (q-1)*3 + 2, (q-1)*3 + 3]
                                    if any(m in meses_26_pres for m in meses_tri):
                                        cols_meses_atual = []
                                        for m_num in meses_tri:
                                            col_m = f"{meses_pt[m_num]}/{ano_str} ({nome_met})"
                                            if col_m in df_final_vol.columns:
                                                cols_finais.append(col_m)
                                                cols_meses_atual.append(col_m)
                                        
                                        nome_meta = f'Meta {q}° Tri ({nome_met})'
                                        df_final_vol[nome_meta] = df_final_vol.get(nome_meta, 0.0).round(2)
                                        cols_finais.append(nome_meta)
                                        
                                        nome_real = f'Real {q}° Tri ({nome_met})'
                                        df_final_vol[nome_real] = df_final_vol[cols_meses_atual].sum(axis=1).round(2) if cols_meses_atual else 0.0
                                        cols_finais.append(nome_real)

                                        nome_pct = f'% Realizado {q}° Tri ({nome_met})'
                                        df_final_vol[nome_pct] = np.where(df_final_vol[nome_meta] > 0, (df_final_vol[nome_real] / df_final_vol[nome_meta]), 0.0)
                                        df_final_vol[nome_pct] = df_final_vol[nome_pct].round(4)
                                        cols_finais.append(nome_pct)

                    df_final_vol = df_final_vol[[c for c in cols_finais if c in df_final_vol.columns]]
                    
                    st.success("Volume processado com sucesso!")
                    st.dataframe(df_final_vol.head())
                    
                    out_vol = io.BytesIO()
                    with pd.ExcelWriter(out_vol, engine="xlsxwriter") as writer:
                        df_final_vol.to_excel(writer, sheet_name="Metas Volume MS", index=False)
                    out_vol.seek(0)
                    st.download_button("📥 Baixar Planilha de Volume", data=out_vol, file_name="Market_Share_Volume.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    st.write("---")
            except Exception as e:
                st.error(f"Erro no processamento de Volume: {e}")

        # ==============================================================
        # PROCESSAMENTO 2: EXECUÇÃO (SKUs, mixRGB, Drop)
        # ==============================================================
        if uploaded_exec is not None:
            try:
                st.markdown("### 📋 Processamento de Execução")
                df_ex = pd.read_csv(uploaded_exec) if uploaded_exec.name.endswith('.csv') else pd.read_excel(uploaded_exec)
                
                mapa_nomes = {
                    'consideraSkuTotal': 'Skus', 
                    'mixRGB': 'mixRGB', 
                    'Drop Por Pdv': 'Drop',
                    'Drop por PDV': 'Drop',
                    'Drop Por PDV': 'Drop'
                }
                df_ex = df_ex.rename(columns=mapa_nomes)
                
                if 'RefMes' not in df_ex.columns or 'CodCli' not in df_ex.columns:
                    st.error("As colunas 'RefMes' e 'CodCli' são obrigatórias na base de Execução.")
                else:
                    df_ex['RefMes'] = pd.to_datetime(df_ex['RefMes'], errors='coerce')
                    df_ex = df_ex.dropna(subset=['RefMes'])
                    df_ex['MesFormatado'] = df_ex['RefMes'].dt.month.map(meses_pt) + '/' + df_ex['RefMes'].dt.strftime('%y')
                    df_ex['Trimestre'] = df_ex['RefMes'].dt.quarter
                    df_ex['Ano'] = df_ex['RefMes'].dt.year
                    
                    metricas_ex = ['Skus', 'mixRGB', 'Drop']
                    for col in metricas_ex:
                        if col in df_ex.columns:
                            df_ex[col] = df_ex[col].astype(str).str.replace(',', '.', regex=False)
                            df_ex[col] = pd.to_numeric(df_ex[col], errors='coerce').fillna(0)
                        else:
                            df_ex[col] = 0.0
                            
                    col_cad_exist_ex = [c for c in colunas_cadastrais if c in df_ex.columns]
                    df_cad_ex = df_ex[['CodCli'] + col_cad_exist_ex].drop_duplicates(subset=['CodCli'], keep='last')
                    
                    # Metas 2025 -> 2026 (Regras Específicas de Execução)
                    df_2025_ex = df_ex[df_ex['Ano'] == 2025].copy()
                    df_metas_ex = pd.DataFrame({'CodCli': df_cad_ex['CodCli'].unique()})
                    
                    if not df_2025_ex.empty:
                        df_tri_25_ex = df_2025_ex.groupby(['CodCli', 'Trimestre'])[metricas_ex].sum().unstack(fill_value=0)
                        df_tri_25_ex.columns = [f"Q{q}_{met}" for met, q in df_tri_25_ex.columns]
                        df_tri_25_ex = df_tri_25_ex.reset_index()
                        for met in metricas_ex:
                            for q in [1, 2, 3, 4]:
                                if f"Q{q}_{met}" not in df_tri_25_ex.columns: df_tri_25_ex[f"Q{q}_{met}"] = 0.0
                                    
                        def calc_meta_ex(row, met):
                            q1, q2, q3, q4 = row[f'Q1_{met}'], row[f'Q2_{met}'], row[f'Q3_{met}'], row[f'Q4_{met}']
                            ultimo = q4 if q4 > 0 else (q3 if q3 > 0 else (q2 if q2 > 0 else (q1 if q1 > 0 else 0)))
                            b1 = q1 if q1 > 0 else ultimo
                            b2 = q2 if q2 > 0 else b1
                            b3 = q3 if q3 > 0 else b2
                            b4 = q4 if q4 > 0 else b3
                            
                            # +1 para Skus e mixRGB, +10% para Drop
                            if met in ['Skus', 'mixRGB']:
                                return pd.Series([b1+1 if b1>0 else 1, b2+1 if b2>0 else 1, b3+1 if b3>0 else 1, b4+1 if b4>0 else 1])
                            else:
                                return pd.Series([b1*1.10, b2*1.10, b3*1.10, b4*1.10])
                            
                        for met in metricas_ex:
                            cols = [f'Meta {q}° Tri ({met})' for q in [1, 2, 3, 4]]
                            df_tri_25_ex[cols] = df_tri_25_ex.apply(lambda r: calc_meta_ex(r, met), axis=1)
                            
                        cols_merge = ['CodCli'] + [f'Meta {q}° Tri ({met})' for met in metricas_ex for q in [1, 2, 3, 4]]
                        df_metas_ex = pd.merge(df_metas_ex, df_tri_25_ex[cols_merge], on='CodCli', how='left').fillna(0)
                    else:
                        for met in metricas_ex:
                            for q in [1, 2, 3, 4]: df_metas_ex[f'Meta {q}° Tri ({met})'] = 0.0

                    # Pivot
                    df_pivot_ex = df_ex.pivot_table(index='CodCli', columns='MesFormatado', values=metricas_ex, aggfunc='sum', fill_value=0)
                    df_pivot_ex.columns = [f"{mes} ({met})" for met, mes in df_pivot_ex.columns]
                    df_pivot_ex = df_pivot_ex.reset_index()
                    
                    df_final_ex = pd.merge(df_cad_ex, df_pivot_ex, on='CodCli', how='left').fillna(0)
                    df_final_ex = pd.merge(df_final_ex, df_metas_ex, on='CodCli', how='left').fillna(0)

                    # Ordenação e Realizado
                    anos_unicos_ex = sorted(df_ex['Ano'].unique())
                    meses_26_pres_ex = df_ex[df_ex['Ano'] == 2026]['RefMes'].dt.month.unique() if 2026 in anos_unicos_ex else []
                    cols_finais_ex = ['CodCli'] + col_cad_exist_ex
                    
                    for met in metricas_ex:
                        for ano in anos_unicos_ex:
                            ano_str = str(ano)[-2:]
                            if ano < 2026:
                                for m_num in range(1, 13):
                                    col_m = f"{meses_pt[m_num]}/{ano_str} ({met})"
                                    if col_m in df_final_ex.columns: cols_finais_ex.append(col_m)
                            elif ano == 2026:
                                for q in [1, 2, 3, 4]:
                                    meses_tri = [(q-1)*3 + 1, (q-1)*3 + 2, (q-1)*3 + 3]
                                    if any(m in meses_26_pres_ex for m in meses_tri):
                                        cols_meses_atual = []
                                        for m_num in meses_tri:
                                            col_m = f"{meses_pt[m_num]}/{ano_str} ({met})"
                                            if col_m in df_final_ex.columns:
                                                cols_finais_ex.append(col_m)
                                                cols_meses_atual.append(col_m)
                                        
                                        nome_meta = f'Meta {q}° Tri ({met})'
                                        df_final_ex[nome_meta] = df_final_ex.get(nome_meta, 0.0).round(2)
                                        cols_finais_ex.append(nome_meta)
                                        
                                        nome_real = f'Real {q}° Tri ({met})'
                                        df_final_ex[nome_real] = df_final_ex[cols_meses_atual].sum(axis=1).round(2) if cols_meses_atual else 0.0
                                        cols_finais_ex.append(nome_real)

                                        nome_pct = f'% Realizado {q}° Tri ({met})'
                                        df_final_ex[nome_pct] = np.where(df_final_ex[nome_meta] > 0, (df_final_ex[nome_real] / df_final_ex[nome_meta]), 0.0)
                                        df_final_ex[nome_pct] = df_final_ex[nome_pct].round(4)
                                        cols_finais_ex.append(nome_pct)

                    df_final_ex = df_final_ex[[c for c in cols_finais_ex if c in df_final_ex.columns]]
                    
                    st.success("Execução processada com sucesso!")
                    st.dataframe(df_final_ex.head())
                    
                    out_ex = io.BytesIO()
                    with pd.ExcelWriter(out_ex, engine="xlsxwriter") as writer:
                        df_final_ex.to_excel(writer, sheet_name="Metas Execução", index=False)
                    out_ex.seek(0)
                    st.download_button("📥 Baixar Planilha de Execução", data=out_ex, file_name="Market_Share_Execucao.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            
            except Exception as e:
                st.error(f"Erro no processamento de Execução: {e}")
                
# ====================================================================
# 7. SETOR DE ASSESSMENT
# ====================================================================
def assessment_page():
    st.title("Setor de Assessment")
    
    col_voltar, col_vazio = st.columns([1, 5])
    with col_voltar:
        if st.button("⬅️ Voltar"):
            st.session_state['current_page'] = 'home'
            st.rerun()

    st.markdown("---")
    
    script_choice = st.selectbox(
        "Selecione uma ferramenta:",
        ("Selecione...", "CMDT", "Controle MPVs"),
        key="assess_select"
    )
    
    # --- FERRAMENTA 1: CMDT ---
    if script_choice == "CMDT":
        st.subheader("Filtro CMDT (Chopeiras e Refrigeradores)")
        st.info("O arquivo deve conter a coluna: **Cmd_Material**")
        
        uploaded_file = st.file_uploader("Envie o arquivo CMDT (.xlsx)", type=["xlsx"], key="cmdt_uploader")
        
        if uploaded_file is not None:
            try:
                df = pd.read_excel(uploaded_file)
                
                coluna_chave = 'Cmd_Material'
                if coluna_chave not in df.columns:
                    st.error(f"Erro: A coluna '{coluna_chave}' não foi encontrada no arquivo.")
                    return

                series_check = df[coluna_chave].astype(str).str.upper().str.strip()

                termos_chopeira = ('CHOPEIRA', 'CHOP', 'CHOPE') 
                mask_chopeira = series_check.str.startswith(termos_chopeira)
                df_chopeiras = df[mask_chopeira].copy()

                termos_refri = ('REF', 'REFR', 'VISA', 'PIL')
                mask_refri = series_check.str.startswith(termos_refri)
                df_refrigeradores = df[mask_refri].copy()

                st.markdown("---")
                c1, c2 = st.columns(2)
                
                with c1:
                    st.success(f"🍺 Chopeiras encontradas: **{len(df_chopeiras)}**")
                    if not df_chopeiras.empty:
                        output_chop = io.BytesIO()
                        with pd.ExcelWriter(output_chop, engine='xlsxwriter') as writer:
                            df_chopeiras.to_excel(writer, index=False)
                        output_chop.seek(0)
                        st.download_button(label="📥 Baixar Chopeiras", data=output_chop, file_name="CMDT_Chopeiras.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                
                with c2:
                    st.success(f"❄️ Refrigeradores encontrados: **{len(df_refrigeradores)}**")
                    if not df_refrigeradores.empty:
                        output_ref = io.BytesIO()
                        with pd.ExcelWriter(output_ref, engine='xlsxwriter') as writer:
                            df_refrigeradores.to_excel(writer, index=False)
                        output_ref.seek(0)
                        st.download_button(label="📥 Baixar Refrigeradores", data=output_ref, file_name="CMDT_Refrigeradores.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

            except Exception as e:
                st.error(f"Erro ao processar o arquivo: {e}")

    # --- FERRAMENTA 2: CONTROLE MPVs ---
    elif script_choice == "Controle MPVs":
        st.subheader("Controle MPVs")
        st.info("Colunas necessárias: **Estoque Atual**, **Estoque Saída**, **Prod (Cód-Descr)**")
        
        uploaded_mpv = st.file_uploader("Envie o arquivo de Controle (.xlsx)", type=["xlsx"], key="mpv_uploader")
        
        if uploaded_mpv is not None:
            try:
                df_mpv = pd.read_excel(uploaded_mpv)
                
                required_cols = ['Estoque Atual', 'Estoque Saída', 'Prod (Cód-Descr)']
                missing_cols = [col for col in required_cols if col not in df_mpv.columns]
                
                if missing_cols:
                    st.error(f"Erro: As seguintes colunas não foram encontradas: {', '.join(missing_cols)}")
                else:
                    total_inicial = len(df_mpv)
                    
                    df_mpv['Estoque Atual'] = pd.to_numeric(df_mpv['Estoque Atual'], errors='coerce').fillna(0)
                    df_mpv = df_mpv[df_mpv['Estoque Atual'] >= 0]
                    
                    df_mpv['Estoque Saída'] = pd.to_numeric(df_mpv['Estoque Saída'], errors='coerce').fillna(0)
                    df_mpv = df_mpv[df_mpv['Estoque Saída'] != 0]
                    
                    palavras_proibidas = [
                        "GARRAFA", "CAIXA", "MESA", "PALETE", "TV", "DIVOSAN", 
                        "REF", "REFR", "CHOPE", "CHOP", "CHOPEIRA"
                    ]
                    
                    col_prod = df_mpv['Prod (Cód-Descr)'].astype(str).str.upper()
                    
                    padrao_regex = r'\b(' + '|'.join(palavras_proibidas) + r')\b'
                    mask_proibidos = col_prod.str.contains(padrao_regex, regex=True, na=False)
                    
                    df_final_mpv = df_mpv[~mask_proibidos].copy()
                    total_final = len(df_final_mpv)
                    
                    st.success(f"Processamento concluído! Linhas restantes: **{total_final}** (de {total_inicial})")
                    st.dataframe(df_final_mpv.head(10))
                    
                    output_mpv = io.BytesIO()
                    with pd.ExcelWriter(output_mpv, engine='xlsxwriter') as writer:
                        df_final_mpv.to_excel(writer, index=False)
                    output_mpv.seek(0)
                    
                    st.download_button(
                        label="📥 Baixar MPVs Filtrados",
                        data=output_mpv,
                        file_name="Controle_MPVs_Filtrado.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

            except Exception as e:
                st.error(f"Erro ao processar MPVs: {e}")

# ====================================================================
# 8. EXECUÇÃO PRINCIPAL
# ====================================================================

if 'is_logged_in' not in st.session_state: st.session_state['is_logged_in'] = False
if 'current_page' not in st.session_state: st.session_state['current_page'] = 'login'
if 'LOGIN_INFO' not in st.session_state: st.session_state['LOGIN_INFO'] = {"admin": "Joao789", "amanda": "12345", "marcia": "54321", "gabi": "G12bi"}

if st.session_state.get('is_logged_in', False):
    page_functions = {
        'home': main_page, 
        'logistics': logistics_page, 
        'commercial': commercial_page,
        'assessment': assessment_page 
    }
    
    current = st.session_state.get('current_page', 'home')
    if current in page_functions:
        page_functions[current]()
    else:
        main_page()
else:
    login_form()
